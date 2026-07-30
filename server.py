#!/usr/bin/env python3
import base64
import html
import json
import mimetypes
import os
import re
import sqlite3
import subprocess
import tempfile
import threading
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


PORT = 4173
ROOT = Path(__file__).resolve().parent
PUBLIC_DIR = ROOT / "public"
SCRIPTS_DIR = ROOT / "scripts"
DATA_DIR = ROOT / "data"
LOCAL_DB_FILE = DATA_DIR / "appexpo_local.db"
SCHEDULER_TRACE_FILE = DATA_DIR / "scheduler_trace.log"
APP_DATA = json.loads((SCRIPTS_DIR / "app_data.json").read_text(encoding="utf-8"))
COUNTRIES = APP_DATA["countries"]
PAGE_TYPES = APP_DATA["pageTypes"]
KNOWN_NAMES = APP_DATA["knownNames"]
STATIC_GAMES_BY_COUNTRY = APP_DATA["staticGames"]
DEVELOPER_MAPPING_FILES = sorted(DATA_DIR.glob("developer_game_mapping_*.json"))
DEVELOPER_MAPPING = {}
if DEVELOPER_MAPPING_FILES:
    try:
        DEVELOPER_MAPPING = json.loads(DEVELOPER_MAPPING_FILES[-1].read_text(encoding="utf-8"))
    except Exception:
        DEVELOPER_MAPPING = {}
EXCEL_GAME_MAPPING_CANDIDATES = [
    ROOT.parent / "AppExpo_游戏分类汇总.xlsx",
    ROOT / "AppExpo_游戏分类汇总.xlsx",
]
MIME_TYPES = {
    ".html": "text/html; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".js": "application/javascript; charset=utf-8",
    ".json": "application/json; charset=utf-8",
    ".png": "image/png",
    ".svg": "image/svg+xml; charset=utf-8",
}
scheduler_lock = threading.Lock()
queue_lock = threading.Lock()
db_lock = threading.Lock()
scheduler_thread = None
weekly_scheduler_thread = None
scheduler_state = {
    "enabled": False,
    "running": False,
    "intervalMinutes": 120,
    "nextRunAt": "",
    "lastRunAt": "",
    "lastFinishedAt": "",
    "lastStatus": "idle",
    "lastError": "",
    "currentGame": "",
    "currentIndex": 0,
    "totalGames": 0,
    "lastRunId": "",
    "lastUploadedGames": 0,
    "lastMatchCount": 0,
    "pendingQueueCount": 0,
    "lastTriggerSource": "",
    "currentTriggerSource": "",
    "progressEvents": [],
    "retryPendingCount": 0,
}
weekly_scheduler_state = {
    "running": False,
    "lastRunAt": "",
    "lastFinishedAt": "",
    "lastStatus": "idle",
    "lastError": "",
    "currentCountry": "",
    "currentIndex": 0,
    "totalCountries": 0,
    "lastInserted": 0,
    "lastUpdated": 0,
}

SYNC_RETRY_ROUNDS = 12
SYNC_RETRY_SLEEP_SECONDS = 4.0
WEEKLY_CAPTURE_RETRY_ROUNDS = 1
WEEKLY_CAPTURE_RETRY_SLEEP_SECONDS = 8.0
WEEKLY_CAPTURE_RUN_HOUR = 6
WEEKLY_CAPTURE_RUN_MINUTE = 0
SCHEDULER_PROGRESS_LIMIT = 240
mapping_name_index_by_id = None
mapping_name_index_by_name = None
merged_static_games_cache = None
excel_game_rules_cache = None
excel_game_rules_by_id = None
excel_game_rules_by_name = None
app_icon_lookup_cache = {}
HISTORY_GAME_PRIORITY_NAMES = [
    "伊瑟 / Etheria: Restart",
    "出发吧麦芬",
    "心动小镇",
    "火炬之光：无限",
    "铃兰之剑",
    "香肠派对",
    "GIRLS' FRONTLINE 2",
    "Metal Slug: Awakening",
    "Reverse: 1999",
]


class AppleCapacityError(Exception):
    pass


class SyncUploadError(Exception):
    pass


def utc_now_iso():
    return datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")


def scheduler_snapshot():
    with scheduler_lock:
        return dict(scheduler_state)


def update_scheduler_state(**kwargs):
    with scheduler_lock:
        scheduler_state.update(kwargs)


def reset_scheduler_progress():
    update_scheduler_state(progressEvents=[], retryPendingCount=0)


def push_scheduler_progress(message, status="info", country="", page_type="", retry=False):
    event = {
        "time": utc_now_iso(),
        "message": clean_text(message, 240),
        "status": status or "info",
        "country": (country or "").upper(),
        "pageType": (page_type or "").lower(),
        "retry": bool(retry),
    }
    with scheduler_lock:
        events = list(scheduler_state.get("progressEvents") or [])
        events.append(event)
        scheduler_state["progressEvents"] = events[-SCHEDULER_PROGRESS_LIMIT:]


def ensure_data_dir():
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def get_db_connection():
    ensure_data_dir()
    connection = sqlite3.connect(LOCAL_DB_FILE)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA journal_mode=WAL")
    connection.execute("PRAGMA synchronous=NORMAL")
    return connection


def init_local_db():
    with db_lock:
        connection = get_db_connection()
        try:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS sync_runs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    run_id TEXT NOT NULL UNIQUE,
                    sync_date TEXT NOT NULL,
                    checked_at TEXT NOT NULL,
                    country TEXT NOT NULL,
                    country_label TEXT DEFAULT '',
                    local_name TEXT DEFAULT '',
                    game_count INTEGER DEFAULT 0,
                    page_count INTEGER DEFAULT 0,
                    match_count INTEGER DEFAULT 0,
                    game_keys TEXT DEFAULT '',
                    game_names TEXT DEFAULT '',
                    status TEXT DEFAULT 'completed',
                    source TEXT DEFAULT 'AppExpo',
                    updated_at TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS placement_matches (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    business_key TEXT NOT NULL UNIQUE,
                    sync_date TEXT NOT NULL,
                    checked_at TEXT NOT NULL,
                    run_id TEXT NOT NULL,
                    game_key TEXT DEFAULT '',
                    game_name TEXT DEFAULT '',
                    app_id TEXT DEFAULT '',
                    country TEXT NOT NULL,
                    country_label TEXT DEFAULT '',
                    local_name TEXT DEFAULT '',
                    page_type TEXT DEFAULT '',
                    page_label TEXT DEFAULT '',
                    group_title TEXT DEFAULT '',
                    section_title TEXT DEFAULT '',
                    placement_title TEXT DEFAULT '',
                    subtitle TEXT DEFAULT '',
                    media_mode TEXT DEFAULT '',
                    updated_at TEXT DEFAULT '',
                    image TEXT DEFAULT '',
                    app_icon TEXT DEFAULT '',
                    raw_match TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    touched_at TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS weekly_games_items (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    business_key TEXT NOT NULL UNIQUE,
                    capture_date TEXT NOT NULL,
                    week_start TEXT NOT NULL,
                    week_end TEXT NOT NULL,
                    checked_at TEXT NOT NULL,
                    country TEXT NOT NULL,
                    country_label TEXT DEFAULT '',
                    local_name TEXT DEFAULT '',
                    item_type TEXT NOT NULL,
                    title TEXT DEFAULT '',
                    subtitle TEXT DEFAULT '',
                    description TEXT DEFAULT '',
                    app_id TEXT DEFAULT '',
                    app_title TEXT DEFAULT '',
                    app_subtitle TEXT DEFAULT '',
                    app_icon TEXT DEFAULT '',
                    image TEXT DEFAULT '',
                    video TEXT DEFAULT '',
                    event_status TEXT DEFAULT '',
                    event_kind TEXT DEFAULT '',
                    event_start_date TEXT DEFAULT '',
                    event_end_date TEXT DEFAULT '',
                    event_requirement TEXT DEFAULT '',
                    position INTEGER DEFAULT 0,
                    module_position INTEGER DEFAULT 0,
                    section_title TEXT DEFAULT '',
                    section_subtitle TEXT DEFAULT '',
                    raw_item TEXT DEFAULT '',
                    created_at TEXT NOT NULL,
                    touched_at TEXT NOT NULL
                )
                """
            )
            connection.execute("CREATE INDEX IF NOT EXISTS idx_runs_sync_date ON sync_runs(sync_date)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_runs_country ON sync_runs(country)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_matches_sync_date ON placement_matches(sync_date)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_matches_country ON placement_matches(country)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_matches_game_key ON placement_matches(game_key)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_matches_app_id ON placement_matches(app_id)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_matches_page_type ON placement_matches(page_type)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_weekly_games_week_country ON weekly_games_items(week_start, country)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_weekly_games_capture_country ON weekly_games_items(capture_date, country)")
            connection.execute("CREATE INDEX IF NOT EXISTS idx_weekly_games_type ON weekly_games_items(item_type)")
            weekly_columns = {row["name"] for row in connection.execute("PRAGMA table_info(weekly_games_items)").fetchall()}
            if "video" not in weekly_columns:
                connection.execute("ALTER TABLE weekly_games_items ADD COLUMN video TEXT DEFAULT ''")
            connection.commit()
        finally:
            connection.close()


def append_scheduler_trace(event, **payload):
    ensure_data_dir()
    entry = {
        "time": utc_now_iso(),
        "event": event,
        **payload,
    }
    with (DATA_DIR / SCHEDULER_TRACE_FILE.name).open("a", encoding="utf-8") as handle:
        handle.write(json.dumps(entry, ensure_ascii=False) + "\n")


def pending_queue_count():
    return 0


def enqueue_pending_sync(payload, error_message=""):
    update_scheduler_state(pendingQueueCount=0)
    return {
        "queuedAt": utc_now_iso(),
        "error": clean_text(error_message, 400),
        "payload": payload,
    }


def load_pending_sync_entries():
    return []


def save_pending_sync_entries(entries):
    update_scheduler_state(pendingQueueCount=0)


def flush_pending_sync_queue():
    update_scheduler_state(pendingQueueCount=0)
    return {"flushed": 0, "remaining": 0}


def run_scheduled_full_sync(trigger_source="unknown"):
    catalog = global_game_catalog()["games"]
    checked_at = utc_now_iso()
    run_id = f"appexpo-sync-{int(time.time())}"
    game_map = {game["key"]: game for game in catalog}
    results_by_game = {game["key"]: [] for game in catalog}
    country_game_map = {}
    for game in catalog:
        for country_code in game["countryCodes"]:
            country_game_map.setdefault(country_code, []).append(game["key"])
    page_types = ["today", "games"]
    fetch_jobs = [
        {"countryCode": country_code, "pageType": page_type}
        for country_code in sorted(country_game_map.keys())
        for page_type in page_types
    ]
    country_total = len(fetch_jobs)
    update_scheduler_state(
        running=True,
        lastRunAt=checked_at,
        lastFinishedAt="",
        lastStatus="running",
        lastError="",
        currentGame="",
        currentIndex=0,
        totalGames=country_total,
        lastRunId=run_id,
        lastUploadedGames=0,
        lastMatchCount=0,
        lastTriggerSource=trigger_source,
        currentTriggerSource=trigger_source,
        retryPendingCount=0,
    )
    reset_scheduler_progress()
    append_scheduler_trace("sync_started", triggerSource=trigger_source, runId=run_id, checkedAt=checked_at)
    push_scheduler_progress("已开始同步任务，正在依次请求 Today / Games 接口。", "info")
    uploaded_games = 0
    match_count = 0
    try:
        fetched_payloads = {}
        failed_jobs = []
        for job_index, job in enumerate(fetch_jobs, 1):
            country_code = job["countryCode"]
            page_type = job["pageType"]
            related_game_keys = country_game_map.get(country_code) or []
            update_scheduler_state(
                currentGame=f"{country_code.upper()} · {PAGE_TYPES[page_type]['label']}",
                currentIndex=job_index,
                totalGames=country_total,
            )
            country = COUNTRIES[country_code]
            url = editorial_url(country_code, page_type)
            push_scheduler_progress(
                f"请求 {country_code.upper()} · {PAGE_TYPES[page_type]['label']} 接口",
                "info",
                country=country_code,
                page_type=page_type,
            )
            try:
                data = fetch_json(url, country, retries=2, editorial=True)
                fetched_payloads[(country_code, page_type)] = data
                push_scheduler_progress(
                    f"{country_code.upper()} · {PAGE_TYPES[page_type]['label']} 请求成功",
                    "success",
                    country=country_code,
                    page_type=page_type,
                )
            except Exception as error:
                error_message = "Apple API 临时限流，请稍后重新分析" if isinstance(error, AppleCapacityError) or re.search(r"429|capacity exceeded", str(error), re.I) else str(error) or "接口请求失败"
                failed_jobs.append({
                    "countryCode": country_code,
                    "pageType": page_type,
                    "url": url,
                    "error": error_message,
                })
                push_scheduler_progress(
                    f"{country_code.upper()} · {PAGE_TYPES[page_type]['label']} 请求失败，已加入收尾重试",
                    "failed",
                    country=country_code,
                    page_type=page_type,
                )
        remaining_failed_jobs = list(failed_jobs)
        if remaining_failed_jobs:
            push_scheduler_progress(
                f"首轮完成，开始补重试 {len(remaining_failed_jobs)} 个失败接口。",
                "retry",
            )
        for retry_round in range(1, SYNC_RETRY_ROUNDS + 1):
            if not remaining_failed_jobs:
                break
            round_failed_jobs = []
            total_retry_jobs = len(remaining_failed_jobs)
            update_scheduler_state(retryPendingCount=total_retry_jobs)
            push_scheduler_progress(
                f"开始第 {retry_round} 轮补重试，剩余 {total_retry_jobs} 个接口。",
                "retry",
            )
            for retry_index, job in enumerate(remaining_failed_jobs, 1):
                country_code = job["countryCode"]
                page_type = job["pageType"]
                country = COUNTRIES[country_code]
                update_scheduler_state(
                    currentGame=f"第 {retry_round} 轮补重试 · {country_code.upper()} · {PAGE_TYPES[page_type]['label']}",
                    currentIndex=retry_index,
                    totalGames=max(1, total_retry_jobs),
                    retryPendingCount=max(0, total_retry_jobs - retry_index + 1),
                )
                push_scheduler_progress(
                    f"第 {retry_round} 轮补重试 {country_code.upper()} · {PAGE_TYPES[page_type]['label']}",
                    "retry",
                    country=country_code,
                    page_type=page_type,
                    retry=True,
                )
                try:
                    data = fetch_json(job["url"], country, retries=2, editorial=True)
                    fetched_payloads[(country_code, page_type)] = data
                    push_scheduler_progress(
                        f"{country_code.upper()} · {PAGE_TYPES[page_type]['label']} 第 {retry_round} 轮补重试成功",
                        "success",
                        country=country_code,
                        page_type=page_type,
                        retry=True,
                    )
                except Exception as error:
                    error_message = "Apple API 临时限流，请稍后重新分析" if isinstance(error, AppleCapacityError) or re.search(r"429|capacity exceeded", str(error), re.I) else str(error) or "接口请求失败"
                    round_failed_jobs.append({
                        "countryCode": country_code,
                        "pageType": page_type,
                        "url": job["url"],
                        "error": error_message,
                    })
                    push_scheduler_progress(
                        f"{country_code.upper()} · {PAGE_TYPES[page_type]['label']} 第 {retry_round} 轮补重试仍失败",
                        "failed",
                        country=country_code,
                        page_type=page_type,
                        retry=True,
                    )
            remaining_failed_jobs = round_failed_jobs
            update_scheduler_state(retryPendingCount=len(remaining_failed_jobs))
            if remaining_failed_jobs and retry_round < SYNC_RETRY_ROUNDS:
                push_scheduler_progress(
                    f"第 {retry_round} 轮结束，仍有 {len(remaining_failed_jobs)} 个接口失败，稍后继续下一轮。",
                    "retry",
                )
                time.sleep(SYNC_RETRY_SLEEP_SECONDS)
        if remaining_failed_jobs:
            failed_summary = "；".join(
                f"{item['countryCode'].upper()}·{PAGE_TYPES[item['pageType']]['label']}"
                for item in remaining_failed_jobs[:6]
            )
            raise RuntimeError(f"仍有 {len(remaining_failed_jobs)} 个接口失败：{failed_summary}")
        for country_code in sorted(country_game_map.keys()):
            related_game_keys = country_game_map.get(country_code) or []
            for page_type in page_types:
                data = fetched_payloads.get((country_code, page_type))
                if data is None:
                    continue
                for game_key in related_game_keys:
                    game = game_map.get(game_key)
                    if not game:
                        continue
                    country_entry = game["countryEntries"].get(country_code) or {}
                    result = analyze_country_page_payload(
                        country_code=country_code,
                        page_type=page_type,
                        data=data,
                        game_name=clean_text(country_entry.get("name"), 120) or game["displayName"],
                        app_id=str(country_entry.get("appId") or game["id"]),
                        aliases=[clean_text(item, 120) for item in (country_entry.get("aliases") or game.get("aliases") or []) if clean_text(item, 120)],
                        checked_at=checked_at,
                        app_icon=str(country_entry.get("icon") or ""),
                    )
                    results_by_game[game_key].append(result)
        for game in catalog:
            results = results_by_game.get(game["key"]) or []
            match_count += sum(len(result.get("matches") or []) for result in results)
        country_payloads = build_country_sync_payloads(run_id, checked_at, catalog, results_by_game)
        push_scheduler_progress(
            f"全部接口成功，开始写入 {len(country_payloads)} 个国家的本地历史记录。",
            "success",
        )
        for index, payload in enumerate(country_payloads, 1):
            country_games = payload.get("games") or []
            country_code = (((country_games[0] or {}).get("results") or [{}])[0].get("country") or "")
            local_name = (((country_games[0] or {}).get("results") or [{}])[0].get("localName") or country_code)
            update_scheduler_state(currentGame=f"写入数据库 · {country_code} · {local_name}", currentIndex=index, totalGames=country_total)
            push_scheduler_progress(
                f"开始写入 {country_code.upper()} · {local_name} 本地记录（{index}/{len(country_payloads)}）",
                "info",
                country=country_code,
            )
            post_sync_payload(payload)
            uploaded_games += 1
            update_scheduler_state(lastUploadedGames=uploaded_games, lastMatchCount=match_count)
            push_scheduler_progress(
                f"已写入 {country_code.upper()} 本地记录（{uploaded_games}/{len(country_payloads)}）",
                "success",
                country=country_code,
            )
        update_scheduler_state(
            running=False,
            lastFinishedAt=utc_now_iso(),
            lastStatus="completed",
            currentGame="",
            currentIndex=0,
            totalGames=country_total,
            lastUploadedGames=uploaded_games,
            lastMatchCount=match_count,
            currentTriggerSource="",
            retryPendingCount=0,
        )
        push_scheduler_progress("全部接口成功，请求结果已写入本地历史记录。", "success")
        append_scheduler_trace(
            "sync_completed",
            triggerSource=trigger_source,
            runId=run_id,
            checkedAt=checked_at,
            uploadedGames=uploaded_games,
            matchCount=match_count,
        )
    except Exception as error:
        update_scheduler_state(
            running=False,
            lastFinishedAt=utc_now_iso(),
            lastStatus="failed",
            lastError=str(error)[:300],
            currentGame="",
            lastUploadedGames=uploaded_games,
            lastMatchCount=match_count,
            currentTriggerSource="",
            retryPendingCount=0,
        )
        push_scheduler_progress(f"同步失败：{str(error)[:180]}", "failed")
        append_scheduler_trace(
            "sync_failed",
            triggerSource=trigger_source,
            runId=run_id,
            checkedAt=checked_at,
            uploadedGames=uploaded_games,
            matchCount=match_count,
            error=str(error)[:300],
        )


def scheduler_loop():
    while True:
        snapshot = scheduler_snapshot()
        if not snapshot["enabled"]:
            time.sleep(1)
            continue
        if snapshot["running"]:
            time.sleep(1)
            continue
        next_run = snapshot["nextRunAt"]
        if not next_run:
            next_ts = time.time() + snapshot["intervalMinutes"] * 60
            update_scheduler_state(nextRunAt=datetime.fromtimestamp(next_ts, timezone.utc).isoformat().replace("+00:00", "Z"))
            time.sleep(1)
            continue
        try:
            next_ts = datetime.fromisoformat(next_run.replace("Z", "+00:00")).timestamp()
        except Exception:
            next_ts = time.time()
        if time.time() < next_ts:
            time.sleep(1)
            continue
        append_scheduler_trace("scheduler_loop_trigger", nextRunAt=next_run, intervalMinutes=snapshot["intervalMinutes"])
        run_scheduled_full_sync("scheduler-loop")
        fresh = scheduler_snapshot()
        if fresh["enabled"]:
            upcoming_ts = time.time() + fresh["intervalMinutes"] * 60
            update_scheduler_state(nextRunAt=datetime.fromtimestamp(upcoming_ts, timezone.utc).isoformat().replace("+00:00", "Z"))
        else:
            update_scheduler_state(nextRunAt="")


def ensure_scheduler_thread():
    global scheduler_thread
    if scheduler_thread and scheduler_thread.is_alive():
        return
    scheduler_thread = threading.Thread(target=scheduler_loop, daemon=True, name="appexpo-scheduler")
    scheduler_thread.start()


def start_scheduler(interval_minutes):
    ensure_scheduler_thread()
    next_ts = time.time() + max(1, interval_minutes) * 60
    update_scheduler_state(
        enabled=True,
        intervalMinutes=max(1, interval_minutes),
        nextRunAt=datetime.fromtimestamp(next_ts, timezone.utc).isoformat().replace("+00:00", "Z"),
        lastStatus="scheduled",
        lastError="",
        currentGame="",
        currentIndex=0,
        retryPendingCount=0,
        progressEvents=[],
    )
    append_scheduler_trace("scheduler_started", intervalMinutes=max(1, interval_minutes))
    return scheduler_snapshot()


def stop_scheduler():
    append_scheduler_trace("scheduler_stopped")
    running = scheduler_snapshot()["running"]
    update_scheduler_state(
        enabled=False,
        nextRunAt="",
        currentGame="",
        currentIndex=0,
        retryPendingCount=0,
        lastStatus="stopped" if not running else scheduler_snapshot()["lastStatus"],
        progressEvents=[] if not running else scheduler_snapshot()["progressEvents"],
    )
    return scheduler_snapshot()


def local_today():
    return datetime.now().date().isoformat()


def week_bounds_for_day(day_text):
    try:
        day = datetime.fromisoformat(str(day_text or "")[:10]).date()
    except Exception:
        day = datetime.now().date()
    week_start = day - timedelta(days=day.weekday())
    week_end = week_start + timedelta(days=6)
    return week_start.isoformat(), week_end.isoformat()


def weekly_snapshot():
    with scheduler_lock:
        return dict(weekly_scheduler_state)


def update_weekly_state(**kwargs):
    with scheduler_lock:
        weekly_scheduler_state.update(kwargs)


def weekly_item_key(row):
    return "|".join([
        row.get("capture_date") or "",
        row.get("country") or "",
        row.get("item_type") or "",
        str(row.get("position") or 0),
        row.get("app_id") or "",
        row.get("title") or "",
    ])


def weekly_country_has_capture(country_code, capture_date):
    rows = sql_fetch_all(
        "SELECT 1 FROM weekly_games_items WHERE country = ? AND capture_date = ? LIMIT 1",
        (country_code, capture_date),
    )
    return bool(rows)


def weekly_capture_count(capture_date):
    rows = sql_fetch_all(
        "SELECT COUNT(DISTINCT country) AS total FROM weekly_games_items WHERE capture_date = ?",
        (capture_date,),
    )
    return int((rows[0] or {}).get("total") or 0) if rows else 0


def weekly_match_to_row(match, country_code, checked_at, capture_date=None):
    capture_date = capture_date or local_today()
    week_start, week_end = week_bounds_for_day(capture_date)
    item_type = "carousel" if match.get("mediaMode") == "carousel" else "event"
    title = clean_text(match.get("heroTitle") or match.get("placementTitle") or match.get("sectionTitle") or "", 220)
    subtitle = clean_text(match.get("heroDescription") or match.get("subtitle") or match.get("sectionSubtitle") or "", 260)
    row = {
        "capture_date": capture_date,
        "week_start": week_start,
        "week_end": week_end,
        "checked_at": checked_at,
        "country": country_code,
        "country_label": COUNTRIES[country_code].get("label") or "",
        "local_name": COUNTRIES[country_code].get("localName") or "",
        "item_type": item_type,
        "title": title,
        "subtitle": subtitle,
        "description": clean_text(match.get("description") or "", 600),
        "app_id": clean_text(match.get("id") or match.get("appId") or "", 80),
        "app_title": clean_text(match.get("appTitle") or "", 220),
        "app_subtitle": clean_text(match.get("appSubtitle") or "", 260),
        "app_icon": match.get("appIcon") or match.get("iconImage") or "",
        "image": match.get("image") or match.get("heroImage") or "",
        "video": match.get("video") or "",
        "event_status": clean_text(match.get("eventStatus") or "", 80),
        "event_kind": clean_text(match.get("eventKind") or match.get("heroEyebrow") or "", 120),
        "event_start_date": clean_text(match.get("eventStartDate") or "", 80),
        "event_end_date": clean_text(match.get("eventEndDate") or "", 80),
        "event_requirement": clean_text(match.get("eventRequirement") or match.get("buttonNote") or "", 160),
        "position": int(match.get("position") or 0),
        "module_position": int(match.get("modulePosition") or 0),
        "section_title": clean_text(match.get("sectionTitle") or "", 220),
        "section_subtitle": clean_text(match.get("sectionSubtitle") or "", 260),
        "raw_item": json.dumps(match, ensure_ascii=False),
    }
    row["business_key"] = weekly_item_key(row)
    return row


def extract_weekly_games_items(country_code, data, checked_at, capture_date=None):
    placements = build_games_placement_index(data, {
        "pageType": "games",
        "gameName": "",
        "appId": "",
        "aliases": [],
        "checkedAt": checked_at,
        "appIcon": "",
    })
    items = []
    seen = set()
    for match in placements:
        if match.get("mediaMode") not in {"carousel", "event"}:
            continue
        if not has_placement_image(match):
            continue
        row = weekly_match_to_row(match, country_code, checked_at, capture_date)
        key = row["business_key"]
        if key in seen:
            continue
        seen.add(key)
        items.append(row)
    items.sort(key=lambda item: (
        0 if item.get("item_type") == "carousel" else 1,
        int(item.get("module_position") or 0),
        int(item.get("position") or 0),
        item.get("title") or "",
    ))
    return items


def store_weekly_games_items(rows):
    if not rows:
        return {"inserted": 0, "updated": 0}
    inserted = 0
    updated = 0
    now = utc_now_iso()
    with db_lock:
        connection = get_db_connection()
        try:
            for row in rows:
                values = (
                    row.get("business_key") or "",
                    row.get("capture_date") or "",
                    row.get("week_start") or "",
                    row.get("week_end") or "",
                    row.get("checked_at") or "",
                    row.get("country") or "",
                    row.get("country_label") or "",
                    row.get("local_name") or "",
                    row.get("item_type") or "",
                    row.get("title") or "",
                    row.get("subtitle") or "",
                    row.get("description") or "",
                    row.get("app_id") or "",
                    row.get("app_title") or "",
                    row.get("app_subtitle") or "",
                    row.get("app_icon") or "",
                    row.get("image") or "",
                    row.get("video") or "",
                    row.get("event_status") or "",
                    row.get("event_kind") or "",
                    row.get("event_start_date") or "",
                    row.get("event_end_date") or "",
                    row.get("event_requirement") or "",
                    int(row.get("position") or 0),
                    int(row.get("module_position") or 0),
                    row.get("section_title") or "",
                    row.get("section_subtitle") or "",
                    row.get("raw_item") or "",
                    now,
                    now,
                )
                before = connection.total_changes
                connection.execute(
                    """
                    INSERT INTO weekly_games_items (
                        business_key, capture_date, week_start, week_end, checked_at,
                        country, country_label, local_name, item_type, title, subtitle,
                        description, app_id, app_title, app_subtitle, app_icon, image, video,
                        event_status, event_kind, event_start_date, event_end_date,
                        event_requirement, position, module_position, section_title,
                        section_subtitle, raw_item, created_at, touched_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(business_key) DO UPDATE SET
                        checked_at = excluded.checked_at,
                        country_label = excluded.country_label,
                        local_name = excluded.local_name,
                        title = excluded.title,
                        subtitle = excluded.subtitle,
                        description = excluded.description,
                        app_title = excluded.app_title,
                        app_subtitle = excluded.app_subtitle,
                        app_icon = excluded.app_icon,
                        image = excluded.image,
                        video = excluded.video,
                        event_status = excluded.event_status,
                        event_kind = excluded.event_kind,
                        event_start_date = excluded.event_start_date,
                        event_end_date = excluded.event_end_date,
                        event_requirement = excluded.event_requirement,
                        module_position = excluded.module_position,
                        section_title = excluded.section_title,
                        section_subtitle = excluded.section_subtitle,
                        raw_item = excluded.raw_item,
                        touched_at = excluded.touched_at
                    """,
                    values,
                )
                if connection.total_changes == before + 1:
                    existing = connection.execute(
                        "SELECT created_at, touched_at FROM weekly_games_items WHERE business_key = ?",
                        (row.get("business_key") or "",),
                    ).fetchone()
                    if existing and existing["created_at"] == existing["touched_at"]:
                        inserted += 1
                    else:
                        updated += 1
            connection.commit()
        finally:
            connection.close()
    return {"inserted": inserted, "updated": updated}


def is_weekly_retryable_error(error):
    return isinstance(error, AppleCapacityError) or re.search(
        r"429|capacity exceeded|IncompleteRead|timed out|connection reset",
        str(error or ""),
        re.I,
    )


def capture_weekly_games_for_countries(country_codes=None, force=False, trigger_source="daily"):
    checked_at = utc_now_iso()
    capture_date = local_today()
    country_codes = [str(code).upper() for code in (country_codes or sorted(COUNTRIES.keys())) if str(code).upper() in COUNTRIES]
    update_weekly_state(
        running=True,
        lastRunAt=checked_at,
        lastFinishedAt="",
        lastStatus="running",
        lastError="",
        currentCountry="",
        currentIndex=0,
        totalCountries=len(country_codes),
        lastInserted=0,
        lastUpdated=0,
    )
    inserted = 0
    updated = 0
    failed = []
    retryable_countries = []
    try:
        for index, country_code in enumerate(country_codes, 1):
            update_weekly_state(currentCountry=country_code, currentIndex=index)
            if not force and weekly_country_has_capture(country_code, capture_date):
                continue
            country = COUNTRIES[country_code]
            try:
                data = fetch_json(editorial_url(country_code, "games"), country, retries=3, editorial=True)
                rows = extract_weekly_games_items(country_code, data, checked_at, capture_date)
                result = store_weekly_games_items(rows)
                inserted += result["inserted"]
                updated += result["updated"]
                update_weekly_state(lastInserted=inserted, lastUpdated=updated)
            except Exception as error:
                message = str(error)[:160]
                if is_weekly_retryable_error(error):
                    retryable_countries.append(country_code)
                else:
                    failed.append(f"{country_code}: {message}")
                update_weekly_state(lastError="；".join(failed[-3:]))
            time.sleep(0.9)
        for retry_round in range(1, WEEKLY_CAPTURE_RETRY_ROUNDS + 1):
            if not retryable_countries:
                break
            time.sleep(WEEKLY_CAPTURE_RETRY_SLEEP_SECONDS)
            pending = retryable_countries
            retryable_countries = []
            update_weekly_state(
                lastError=f"限流补抓第 {retry_round} 轮，剩余 {len(pending)} 个国家",
                totalCountries=len(pending),
                currentIndex=0,
            )
            for retry_index, country_code in enumerate(pending, 1):
                update_weekly_state(currentCountry=country_code, currentIndex=retry_index)
                if not force and weekly_country_has_capture(country_code, capture_date):
                    continue
                country = COUNTRIES[country_code]
                try:
                    data = fetch_json(editorial_url(country_code, "games"), country, retries=3, editorial=True)
                    rows = extract_weekly_games_items(country_code, data, checked_at, capture_date)
                    result = store_weekly_games_items(rows)
                    inserted += result["inserted"]
                    updated += result["updated"]
                    update_weekly_state(lastInserted=inserted, lastUpdated=updated)
                except Exception as error:
                    message = str(error)[:160]
                    if is_weekly_retryable_error(error) and retry_round < WEEKLY_CAPTURE_RETRY_ROUNDS:
                        retryable_countries.append(country_code)
                    else:
                        failed.append(f"{country_code}: {message}")
                    update_weekly_state(lastError="；".join(failed[-3:]) or f"限流补抓第 {retry_round} 轮继续等待")
                time.sleep(0.9)
        for country_code in retryable_countries:
            failed.append(f"{country_code}: Apple API 限流，补抓后仍未完成")
        status = "completed" if not failed else "partial"
        update_weekly_state(
            running=False,
            lastFinishedAt=utc_now_iso(),
            lastStatus=status,
            lastError="；".join(failed[:6]),
            currentCountry="",
            currentIndex=0,
            totalCountries=len(country_codes),
            lastInserted=inserted,
            lastUpdated=updated,
        )
        return {
            "inserted": inserted,
            "updated": updated,
            "checkedAt": checked_at,
            "countryCount": len(country_codes),
            "failed": failed,
            "status": status,
        }
    except Exception as error:
        update_weekly_state(
            running=False,
            lastFinishedAt=utc_now_iso(),
            lastStatus="failed",
            lastError=str(error)[:300],
            currentCountry="",
        )
        raise


def weekly_seconds_until_next_run():
    now = datetime.now()
    today_start = now.replace(hour=WEEKLY_CAPTURE_RUN_HOUR, minute=WEEKLY_CAPTURE_RUN_MINUTE, second=0, microsecond=0)
    if weekly_capture_count(local_today()) >= len(COUNTRIES):
        target = today_start + timedelta(days=1)
    elif now < today_start:
        target = today_start
    else:
        target = (now + timedelta(hours=1)).replace(minute=0, second=0, microsecond=0)
    return max(60, int((target - now).total_seconds()))


def weekly_scheduler_loop():
    time.sleep(3)
    while True:
        try:
            now = datetime.now()
            today_start = now.replace(hour=WEEKLY_CAPTURE_RUN_HOUR, minute=WEEKLY_CAPTURE_RUN_MINUTE, second=0, microsecond=0)
            if now >= today_start and not weekly_snapshot()["running"] and weekly_capture_count(local_today()) < len(COUNTRIES):
                capture_weekly_games_for_countries(trigger_source="startup-or-daily")
        except Exception:
            pass
        time.sleep(weekly_seconds_until_next_run())


def ensure_weekly_scheduler_thread():
    global weekly_scheduler_thread
    if weekly_scheduler_thread and weekly_scheduler_thread.is_alive():
        return
    weekly_scheduler_thread = threading.Thread(target=weekly_scheduler_loop, daemon=True, name="appexpo-weekly-games")
    weekly_scheduler_thread.start()


def load_weekly_games(country_code="CN", week_start=""):
    country_code = clean_text(country_code or "CN", 20).upper()
    if country_code not in COUNTRIES:
        country_code = "CN"
    if not week_start:
        week_start, _week_end = week_bounds_for_day(local_today())
    week_start, week_end = week_bounds_for_day(week_start)
    rows = sql_fetch_all(
        """
        SELECT * FROM weekly_games_items
        WHERE country = ? AND week_start = ?
        ORDER BY capture_date ASC, item_type ASC, module_position ASC, position ASC, id ASC
        """,
        (country_code, week_start),
    )
    day_counts = {
        (datetime.fromisoformat(week_start).date() + timedelta(days=index)).isoformat(): {"carousel": 0, "event": 0}
        for index in range(7)
    }
    carousel = []
    events = []
    for row in rows:
        capture_date = row.get("capture_date") or ""
        item_type = row.get("item_type") or ""
        if capture_date in day_counts and item_type in day_counts[capture_date]:
            day_counts[capture_date][item_type] += 1
        item = {
            "id": row.get("id"),
            "captureDate": capture_date,
            "checkedAt": row.get("checked_at") or "",
            "country": row.get("country") or "",
            "countryLabel": row.get("country_label") or "",
            "localName": row.get("local_name") or "",
            "type": item_type,
            "title": row.get("title") or "",
            "subtitle": row.get("subtitle") or "",
            "description": row.get("description") or "",
            "appId": row.get("app_id") or "",
            "appTitle": row.get("app_title") or "",
            "appSubtitle": row.get("app_subtitle") or "",
            "appIcon": row.get("app_icon") or "",
            "image": row.get("image") or "",
            "video": row.get("video") or "",
            "eventStatus": row.get("event_status") or "",
            "eventKind": row.get("event_kind") or "",
            "eventStartDate": row.get("event_start_date") or "",
            "eventEndDate": row.get("event_end_date") or "",
            "eventRequirement": row.get("event_requirement") or "",
            "position": row.get("position") or 0,
            "sectionTitle": row.get("section_title") or "",
            "sectionSubtitle": row.get("section_subtitle") or "",
        }
        if item_type == "carousel":
            carousel.append(item)
        elif item_type == "event":
            events.append(item)
    days = []
    for index in range(7):
        day = (datetime.fromisoformat(week_start).date() + timedelta(days=index)).isoformat()
        days.append({
            "date": day,
            "weekday": ["周一", "周二", "周三", "周四", "周五", "周六", "周日"][index],
            "carouselCount": day_counts[day]["carousel"],
            "eventCount": day_counts[day]["event"],
        })
    return {
        "country": country_code,
        "countryLabel": COUNTRIES[country_code].get("label") or "",
        "localName": COUNTRIES[country_code].get("localName") or "",
        "weekStart": week_start,
        "weekEnd": week_end,
        "days": days,
        "carousel": carousel,
        "events": events,
        "summary": {
            "carouselCount": len(carousel),
            "eventCount": len(events),
            "captureDayCount": sum(1 for day in days if day["carouselCount"] or day["eventCount"]),
        },
        "scheduler": weekly_snapshot(),
    }


def load_weekly_weeks(country_code="CN"):
    country_code = clean_text(country_code or "CN", 20).upper()
    params = []
    where = ""
    if country_code in COUNTRIES:
        where = "WHERE country = ?"
        params.append(country_code)
    rows = sql_fetch_all(
        f"""
        SELECT week_start, week_end, COUNT(*) AS item_count, COUNT(DISTINCT capture_date) AS day_count
        FROM weekly_games_items
        {where}
        GROUP BY week_start, week_end
        ORDER BY week_start DESC
        """,
        tuple(params),
    )
    return {
        "country": country_code if country_code in COUNTRIES else "",
        "weeks": [
            {
                "weekStart": row.get("week_start") or "",
                "weekEnd": row.get("week_end") or "",
                "itemCount": int(row.get("item_count") or 0),
                "dayCount": int(row.get("day_count") or 0),
            }
            for row in rows
            if row.get("week_start")
        ],
    }


def clean_text(value, limit=180):
    text = html.unescape(str(value or ""))
    text = re.sub(r"<br\s*/?>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text[:limit]


def normalize(value):
    return re.sub(r"[\s'\"’“”`~!！?？.,，。:：;；\-_/\|()[\]{}【】<>《》+*=#&]+", "", str(value or "").lower())


def history_game_priority_index(name):
    normalized_name = normalize(name)
    for index, priority_name in enumerate(HISTORY_GAME_PRIORITY_NAMES):
        normalized_priority = normalize(priority_name)
        if normalized_name == normalized_priority or normalized_name.startswith(normalized_priority) or normalized_priority in normalized_name:
            return index
    return len(HISTORY_GAME_PRIORITY_NAMES)


def has_cjk(value):
    return bool(re.search(r"[\u3400-\u9fff]", str(value or "")))


def uniq_strings(values):
    seen = set()
    output = []
    for value in values:
        text = clean_text(value, 120)
        key = normalize(text)
        if not text or key in seen:
            continue
        seen.add(key)
        output.append(text)
    return output


def uniq_app_ids(values):
    seen = set()
    output = []
    for value in values:
        text = re.sub(r"\D+", "", str(value or ""))
        if not text or text in seen:
            continue
        seen.add(text)
        output.append(text)
    return output


def excel_game_mapping_file():
    for path in EXCEL_GAME_MAPPING_CANDIDATES:
        if path.exists():
            return path
    return None


def split_excel_game_names(value):
    text = clean_text(value, 240)
    if not text:
        return []
    parts = [text]
    if "/" in text:
        parts.extend(piece.strip() for piece in re.split(r"\s*/\s*", text) if piece.strip())
    return uniq_strings(parts)


def split_excel_app_ids(*values):
    raw_parts = []
    for value in values:
        if value is None:
            continue
        for piece in re.split(r"[、，,\s]+", str(value).strip()):
            if piece:
                raw_parts.append(piece)
    return uniq_app_ids(raw_parts)


def extract_cjk_display_name(value):
    text = clean_text(value, 240)
    if not text or not has_cjk(text):
        return ""
    extracted = "".join(re.findall(r"[\u3400-\u9fff0-9A-Za-z]+", text))
    # 只保留连续中文段，避免把 ICEY 这类英文再混回中文主名
    cjk_only = "".join(re.findall(r"[\u3400-\u9fff0-9]+", extracted))
    cjk_only = clean_text(cjk_only, 120)
    return cjk_only if has_cjk(cjk_only) else ""


def excel_canonical_chinese_name(aliases):
    for alias in aliases:
        extracted = extract_cjk_display_name(alias)
        if extracted:
            return extracted
    for alias in aliases:
        if has_cjk(alias):
            return alias
    return ""


def load_excel_game_rules():
    global excel_game_rules_cache, excel_game_rules_by_id, excel_game_rules_by_name
    if excel_game_rules_cache is not None and excel_game_rules_by_id is not None and excel_game_rules_by_name is not None:
        return excel_game_rules_cache

    excel_game_rules_cache = []
    excel_game_rules_by_id = {}
    excel_game_rules_by_name = {}

    excel_file = excel_game_mapping_file()
    if not excel_file or load_workbook is None:
        return excel_game_rules_cache

    try:
        workbook = load_workbook(excel_file, data_only=True)
    except Exception:
        return excel_game_rules_cache

    merged_rules = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            raw_name = row[0]
            cn_id = row[1] if len(row) > 1 else ""
            overseas_id = row[2] if len(row) > 2 else ""
            aliases = split_excel_game_names(raw_name)
            app_ids = split_excel_app_ids(cn_id, overseas_id)
            if not aliases or not app_ids:
                continue

            chinese_name = excel_canonical_chinese_name(aliases)
            display_name = chinese_name or aliases[0]
            normalized_aliases = {normalize(alias) for alias in aliases if normalize(alias)}

            target_rule = None
            for existing in merged_rules:
                if set(existing["appIds"]) & set(app_ids):
                    target_rule = existing
                    break
                if chinese_name and existing.get("chineseName") and normalize(existing["chineseName"]) == normalize(chinese_name):
                    target_rule = existing
                    break
                if normalized_aliases & {normalize(alias) for alias in existing.get("aliases", [])}:
                    target_rule = existing
                    break

            if not target_rule:
                target_rule = {
                    "key": "",
                    "displayName": display_name,
                    "chineseName": chinese_name,
                    "aliases": [],
                    "appIds": [],
                    "sheets": [],
                }
                merged_rules.append(target_rule)

            if chinese_name and not target_rule.get("chineseName"):
                target_rule["chineseName"] = chinese_name
            if has_cjk(display_name) and (not target_rule.get("displayName") or not has_cjk(target_rule["displayName"])):
                target_rule["displayName"] = display_name
            if not target_rule.get("displayName"):
                target_rule["displayName"] = display_name

            target_rule["aliases"] = uniq_strings([*target_rule.get("aliases", []), *aliases])
            target_rule["appIds"] = uniq_app_ids([*target_rule.get("appIds", []), *app_ids])
            target_rule["sheets"] = uniq_strings([*target_rule.get("sheets", []), worksheet.title])

    normalized_rules = []
    for rule in merged_rules:
        key_seed = rule.get("chineseName") or rule.get("displayName") or (rule.get("aliases") or [""])[0]
        rule["key"] = f"xlsx:{normalize(key_seed)}"
        normalized_rules.append(rule)
        for app_id in rule.get("appIds", []):
            excel_game_rules_by_id[app_id] = rule
        for alias in rule.get("aliases", []):
            alias_key = normalize(alias)
            if alias_key:
                excel_game_rules_by_name[alias_key] = rule

    excel_game_rules_cache = normalized_rules
    return excel_game_rules_cache


def excel_game_group_rule(game_id, name="", chinese_name=""):
    load_excel_game_rules()
    app_id = str(game_id or "").strip()
    if app_id and app_id in (excel_game_rules_by_id or {}):
        return excel_game_rules_by_id[app_id]
    for value in [name, chinese_name]:
        key = normalize(value)
        if key and key in (excel_game_rules_by_name or {}):
            return excel_game_rules_by_name[key]
    return None


def ensure_mapping_name_index():
    global mapping_name_index_by_id, mapping_name_index_by_name
    if mapping_name_index_by_id is not None and mapping_name_index_by_name is not None:
        return
    mapping_name_index_by_id = {}
    mapping_name_index_by_name = {}
    for app in (DEVELOPER_MAPPING.get("apps") or []):
        app_id = str(app.get("id") or "").strip()
        preferred_name = clean_text(app.get("preferredName"), 120)
        if app_id and preferred_name and has_cjk(preferred_name):
            mapping_name_index_by_id[app_id] = preferred_name
        for country_names in (app.get("countries") or {}).values():
            for country_name in (country_names or []):
                text = clean_text(country_name, 120)
                key = normalize(text)
                if key and preferred_name and has_cjk(preferred_name):
                    mapping_name_index_by_name[key] = preferred_name


def chinese_name_for_game(game_id, name):
    excel_rule = excel_game_group_rule(game_id, name, name)
    if excel_rule and excel_rule.get("chineseName"):
        return excel_rule["chineseName"]
    ensure_mapping_name_index()
    return (
        KNOWN_NAMES.get(str(game_id))
        or KNOWN_NAMES.get(normalize(name))
        or (mapping_name_index_by_id or {}).get(str(game_id))
        or (mapping_name_index_by_name or {}).get(normalize(name))
        or ""
    )


def game_display_name(name, chinese_name):
    if not chinese_name or normalize(chinese_name) == normalize(name):
        return name
    if has_cjk(name):
        return name
    return f"{name} / {chinese_name}"


SPECIAL_GAME_GROUP_RULES = [
    {
        "key": "merge:t3",
        "ids": {"1576661186", "1602814337"},
        "names": {"火力苏打（t3）", "t3 arena", "t3 アリーナ"},
        "displayName": "火力苏打（T3）",
        "chineseName": "火力苏打（T3）",
        "aliases": ["火力苏打（T3）", "T3 Arena", "T3 アリーナ"],
    },
]


def special_game_group_rule(game_id, name="", chinese_name=""):
    normalized_names = {normalize(name), normalize(chinese_name)}
    for rule in SPECIAL_GAME_GROUP_RULES:
        if str(game_id or "").strip() in rule["ids"]:
            return rule
        if any(item and item in rule["names"] for item in normalized_names):
            return rule
    return None


def preferred_country_name(names, country_code, preferred_name):
    cleaned = uniq_strings(names or [])
    if not cleaned:
        return preferred_name or ""
    if country_code in {"CN", "HK", "TW"}:
        return next((name for name in cleaned if has_cjk(name)), cleaned[0])
    return next((name for name in cleaned if not has_cjk(name)), cleaned[0])


def merged_static_games_by_country():
    global merged_static_games_cache
    if merged_static_games_cache is not None:
        return merged_static_games_cache

    merged = {
        country_code: [list(row) for row in rows]
        for country_code, rows in STATIC_GAMES_BY_COUNTRY.items()
    }
    seen = {
        country_code: {
            (str((row or [""])[0]), normalize((row or ["", ""])[1] if len(row) > 1 else ""))
            for row in rows
            if row and len(row) > 1
        }
        for country_code, rows in merged.items()
    }

    for app in (DEVELOPER_MAPPING.get("apps") or []):
        app_id = str(app.get("id") or "").strip()
        preferred_name = clean_text(app.get("preferredName"), 120)
        primary_english_name = clean_text(app.get("primaryEnglishName"), 120)
        countries = app.get("countries") or {}
        for country_code, names in countries.items():
            if country_code not in COUNTRIES:
                continue
            merged.setdefault(country_code, [])
            seen.setdefault(country_code, set())
            local_name = preferred_country_name(names, country_code, preferred_name or primary_english_name)
            chinese_name = preferred_name if has_cjk(preferred_name) else chinese_name_for_game(app_id, local_name)
            row_key = (app_id, normalize(local_name))
            if not app_id or not local_name or row_key in seen[country_code]:
                continue
            merged[country_code].append([app_id, local_name, chinese_name])
            seen[country_code].add(row_key)

    # 只保留 Excel 白名单中的游戏；表外数据不再进入系统目录
    load_excel_game_rules()
    if excel_game_rules_cache:
        filtered = {}
        for country_code, rows in merged.items():
            kept_rows = []
            seen_keys = set()
            for row in rows:
                if not row:
                    continue
                game_id = str((row[0] if len(row) > 0 else "") or "").strip()
                name = clean_text(row[1] if len(row) > 1 else "", 120)
                chinese_name = clean_text(row[2] if len(row) > 2 else "", 120)
                rule = excel_game_group_rule(game_id, name, chinese_name)
                if not rule:
                    continue
                dedupe_key = (
                    game_id,
                    normalize(rule.get("key") or ""),
                    normalize(chinese_name or name),
                )
                if dedupe_key in seen_keys:
                    continue
                seen_keys.add(dedupe_key)
                kept_rows.append([game_id, name, chinese_name])
            filtered[country_code] = kept_rows
        merged = filtered

    merged_static_games_cache = merged
    return merged_static_games_cache


def static_developer_games(country_code):
    country = COUNTRIES[country_code]
    rows = merged_static_games_by_country().get(country_code, [])
    games = []
    for row in rows:
        game_id, name, chinese_name = row
        final_chinese_name = chinese_name or chinese_name_for_game(game_id, name)
        games.append({
            "id": str(game_id),
            "name": name,
            "chineseName": final_chinese_name,
            "displayName": game_display_name(name, final_chinese_name),
            "aliases": uniq_strings([name, final_chinese_name]),
            "artistName": "XD Entertainment",
            "icon": "",
            "genres": [],
            "primaryGenreName": "Games",
            "url": f"https://apps.apple.com/{country['path']}/app/id{game_id}",
        })
    games.sort(key=lambda item: (
        0 if has_cjk(item["displayName"] or item["name"]) else 1,
        (item["displayName"] or item["name"] or "").lower(),
        item["id"],
    ))
    return {
        "country": country_code,
        "countryLabel": country["label"],
        "localName": country["localName"],
        "source": "内置静态游戏清单",
        "static": True,
        "cached": True,
        "games": games,
    }


def global_game_catalog():
    id_country_counts = {}
    merged_rows = merged_static_games_by_country()
    for country_code, rows in merged_rows.items():
        for game_id, _name, _chinese_name in rows:
            id_country_counts[str(game_id)] = id_country_counts.get(str(game_id), 0) + 1

    grouped = {}
    for country_code, rows in merged_rows.items():
        for game_id, name, chinese_name in rows:
            game_id = str(game_id)
            final_chinese_name = chinese_name or chinese_name_for_game(game_id, name) or (name if has_cjk(name) else "")
            canonical_name = final_chinese_name or name
            excel_rule = excel_game_group_rule(game_id, name, final_chinese_name)
            special_rule = special_game_group_rule(game_id, name, final_chinese_name)
            active_rule = special_rule or excel_rule
            if active_rule:
                group_key = active_rule["key"]
            elif final_chinese_name:
                group_key = f"cn:{normalize(final_chinese_name)}"
            elif id_country_counts.get(game_id, 0) > 1:
                group_key = f"id:{game_id}"
            else:
                group_key = f"name:{normalize(name)}"

            entry = grouped.setdefault(group_key, {
                "key": group_key,
                "id": game_id,
                "displayName": (active_rule.get("displayName") if active_rule else canonical_name),
                "chineseName": (active_rule.get("chineseName") if active_rule else (final_chinese_name or (name if has_cjk(name) else ""))),
                "aliases": set(),
                "appIds": set(),
                "countryEntries": {},
            })

            if final_chinese_name and not entry["chineseName"]:
                entry["chineseName"] = final_chinese_name
            if has_cjk(canonical_name) and (not entry["displayName"] or not has_cjk(entry["displayName"])):
                entry["displayName"] = canonical_name
            if not entry["displayName"]:
                entry["displayName"] = canonical_name

            aliases = uniq_strings([name, final_chinese_name, canonical_name, *(((active_rule or {}).get("aliases")) or [])])
            entry["aliases"].update(aliases)
            entry["appIds"].update(uniq_app_ids([game_id, *(((active_rule or {}).get("appIds")) or [])]))
            entry["countryEntries"][country_code] = {
                "appId": game_id,
                "name": name,
                "displayName": game_display_name(name, final_chinese_name),
                "chineseName": final_chinese_name,
                "aliases": aliases,
                "icon": "",
            }

    games = []
    for entry in grouped.values():
        country_codes = sorted(entry["countryEntries"].keys())
        display_name = entry["chineseName"] or entry["displayName"]
        games.append({
            "key": entry["key"],
            "id": entry["id"],
            "displayName": display_name,
            "chineseName": entry["chineseName"],
            "aliases": uniq_strings(entry["aliases"]),
            "appIds": uniq_app_ids(entry["appIds"]),
            "countryCodes": country_codes,
            "countryEntries": entry["countryEntries"],
            "countryCount": len(country_codes),
        })

    games.sort(key=lambda item: (
        -item["countryCount"],
        0 if has_cjk(item["displayName"]) else 1,
        item["displayName"].lower(),
        item["id"],
    ))
    return {
        "source": "内置静态游戏清单",
        "cached": True,
        "games": games,
        "groupCount": len(games),
        "rawAppCount": len(id_country_counts),
    }


def safe_download_name(value):
    text = clean_text(value or "AppExpo", 140)
    text = re.sub(r'[\\/:*?"<>|\s]+', "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or "AppExpo"


def safe_ascii_download_name(value, fallback="AppExpo"):
    text = safe_download_name(value or fallback)
    text = re.sub(r"[^\x20-\x7E]+", "-", text)
    text = re.sub(r"-+", "-", text).strip("-")
    return text or fallback


def content_disposition(filename):
    ascii_name = safe_ascii_download_name(filename, "AppExpo.png")
    ascii_name = re.sub(r'["\\;]+', "-", ascii_name)
    return f'attachment; filename="{ascii_name}"'


def sync_target_config():
    return {
        "enabled": True,
        "mode": "local-sqlite",
        "path": str(LOCAL_DB_FILE),
        "pendingQueueCount": 0,
    }


def build_sync_headers():
    headers = {
        "content-type": "application/json; charset=utf-8",
        "accept": "application/json,text/plain,*/*",
        "user-agent": "AppExpoSync/1.0",
    }
    if SYNC_UPLOAD_TOKEN:
        headers["authorization"] = f"Bearer {SYNC_UPLOAD_TOKEN}"
    return headers


def coze_enabled():
    return bool(COZE_TOKEN and COZE_WORKSPACE_ID and COZE_RUNS_DB_ID and COZE_MATCHES_DB_ID)


def coze_headers():
    return {
        "content-type": "application/json; charset=utf-8",
        "accept": "application/json,text/plain,*/*",
        "authorization": f"Bearer {COZE_TOKEN}",
        "user-agent": "AppExpoCozeSync/1.0",
    }


def is_retryable_coze_status(status_code):
    return status_code in (429, 500, 502, 503, 504)


def is_retryable_coze_code(code):
    return code in (5000,)


def coze_retry_delay(attempt):
    delays = [1.5, 3.0, 5.0, 8.0, 12.0, 16.0]
    return delays[min(attempt, len(delays) - 1)]


def coze_request(path, payload=None, method="POST"):
    if not COZE_TOKEN:
        raise SyncUploadError("尚未配置 Coze Token，请设置 APPEXPO_COZE_TOKEN")
    url = f"{COZE_API_BASE.rstrip('/')}/{path.lstrip('/')}"
    body_bytes = None if payload is None else json.dumps(payload, ensure_ascii=False).encode("utf-8")
    last_error = None
    for attempt in range(6):
        request = urllib.request.Request(
            url,
            data=body_bytes,
            headers=coze_headers(),
            method=method,
        )
        try:
            with urllib.request.urlopen(request, timeout=SYNC_UPLOAD_TIMEOUT) as response:
                raw = response.read().decode("utf-8", "replace")
                data = json.loads(raw or "{}")
                if data.get("code") not in (0, None):
                    code = data.get("code")
                    message = data.get("msg") or f"Coze API error {code}"
                    if attempt < 5 and is_retryable_coze_code(code):
                        time.sleep(coze_retry_delay(attempt))
                        continue
                    raise SyncUploadError(message)
                return data
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", "replace")
            last_error = SyncUploadError(f"Coze HTTP {error.code}: {body[:300]}")
            if attempt < 5 and is_retryable_coze_status(error.code):
                time.sleep(coze_retry_delay(attempt))
                continue
            raise last_error
        except urllib.error.URLError as error:
            last_error = SyncUploadError(str(error) or "Coze 请求失败")
            if attempt < 5:
                time.sleep(coze_retry_delay(attempt))
                continue
            raise last_error
        except json.JSONDecodeError:
            last_error = SyncUploadError("Coze 返回数据不是有效 JSON")
            if attempt < 5:
                time.sleep(coze_retry_delay(attempt))
                continue
            raise last_error
    raise last_error or SyncUploadError("Coze 请求失败")


def coze_insert_rows(database_id, rows):
    if not rows:
        return {"affected_rows": 0}
    affected_rows = 0
    for row in rows:
        response = coze_request(f"databases/{database_id}/records", {
            "insert_rows": [{key: value for key, value in row.items() if value is not None}],
            "is_async": False,
            "connector_id": COZE_CONNECTOR_ID or "1024",
        })
        affected_rows += int(((response.get("data") or {}).get("affected_rows")) or 0)
    return {"affected_rows": affected_rows}


def coze_query_rows(database_id, field_names=None, page_num=1, page_size=50, filter_payload=None, order_by=None):
    payload = {
        "page_num": page_num,
        "page_size": page_size,
        "is_async": False,
        "connector_id": COZE_CONNECTOR_ID or "1024",
    }
    if field_names:
        payload["select_fields"] = {"field_names": field_names, "is_distinct": False}
    if filter_payload:
        payload["filter"] = filter_payload
    if order_by:
        payload["order_by"] = order_by
    response = coze_request(f"databases/{database_id}/records/query", payload)
    return (response.get("data") or {})


def coze_update_rows(database_id, row, filter_payload):
    update_fields = [{"field_name": key, "value": value} for key, value in row.items() if value is not None]
    response = coze_request(f"databases/{database_id}/records", {
        "update_fields": update_fields,
        "filter": filter_payload,
        "is_async": False,
        "connector_id": COZE_CONNECTOR_ID or "1024",
    }, method="PUT")
    return (response.get("data") or {})


def coze_delete_rows(database_id, filter_payload):
    response = coze_request(f"databases/{database_id}/records", {
        "filter": filter_payload,
        "is_async": False,
        "connector_id": COZE_CONNECTOR_ID or "1024",
    }, method="DELETE")
    return (response.get("data") or {})


def coze_delete_all_rows(database_id, filter_payload, page_size=100):
    total_deleted = 0
    while True:
        queried = coze_query_rows(
            database_id,
            field_names=["id"],
            page_num=1,
            page_size=min(max(page_size, 1), 100),
            filter_payload=filter_payload,
        )
        items = queried.get("items") or []
        if not items:
            break
        for item in items:
            row_id = clean_text(item.get("id") or "", 80)
            if not row_id:
                continue
            id_filter = row_filter([
                {"left": "id", "operation": "equal", "right": row_id},
            ])
            coze_delete_rows(database_id, id_filter)
            total_deleted += 1
            time.sleep(0.05)
        if len(items) < min(max(page_size, 1), 100):
            break
    return {"deleted": total_deleted}


def chunked_rows(rows, size):
    step = max(1, int(size or 1))
    for start in range(0, len(rows), step):
        yield rows[start:start + step]


def checked_day(value):
    text = str(value or "").strip()
    return text[:10] if len(text) >= 10 else text


def row_filter(conditions):
    return {"logic": "and", "conditions": [item for item in conditions if item]}


def run_row_key(row):
    return "|".join([
        row.get("country") or "",
        checked_day(row.get("checked_at")),
    ])


def match_row_key(row):
    return "|".join([
        row.get("country") or "",
        row.get("page_type") or "",
        row.get("app_id") or row.get("game_key") or "",
        row.get("media_mode") or "",
        row.get("section_title") or "",
        row.get("placement_title") or "",
        checked_day(row.get("checked_at")),
    ])


def match_business_key(row):
    return "|".join([
        row.get("country") or "",
        row.get("page_type") or "",
        row.get("app_id") or row.get("game_key") or "",
        row.get("media_mode") or "",
        row.get("section_title") or "",
        row.get("placement_title") or "",
        checked_day(row.get("checked_at")),
    ])


def row_checked_at(row):
    return clean_text(row.get("checked_at") or "", 64)


def row_id_value(row):
    return clean_text(row.get("id") or "", 80)


def delete_rows_by_ids(database_id, rows):
    deleted = 0
    for row in rows or []:
        row_id = row_id_value(row)
        if not row_id:
            continue
        id_filter = row_filter([
            {"left": "id", "operation": "equal", "right": row_id},
        ])
        coze_delete_rows(database_id, id_filter)
        deleted += 1
        time.sleep(0.05)
    return deleted


def choose_preferred_match_row(current_row, candidate_row):
    def score(row):
        raw_match = row.get("raw_match") or ""
        raw_path = ""
        module_position = 9999
        item_position = 9999
        position = 9999
        if raw_match:
            try:
                raw = json.loads(raw_match) or {}
                raw_path = clean_text(raw.get("path") or "", 240)
                module_position = int(raw.get("modulePosition") or 9999)
                item_position = int(raw.get("itemPosition") or raw.get("position") or 9999)
                position = int(raw.get("position") or 9999)
            except Exception:
                raw_path = ""
        return (
            -module_position,
            -item_position,
            -position,
            1 if row.get("group_title") else 0,
            1 if row.get("section_title") else 0,
            1 if row.get("subtitle") else 0,
            1 if row.get("image") else 0,
            1 if raw_path else 0,
            len(raw_match),
        )

    return candidate_row if score(candidate_row) > score(current_row) else current_row


def run_row_filter(row):
    return row_filter([
        {"left": "country", "operation": "equal", "right": row.get("country") or ""},
        {"left": "checked_at", "operation": "like", "right": f"{checked_day(row.get('checked_at'))}%"},
    ])


def match_row_filter(row):
    return row_filter([
        {"left": "game_key", "operation": "equal", "right": row.get("game_key") or ""},
        {"left": "country", "operation": "equal", "right": row.get("country") or ""},
        {"left": "page_type", "operation": "equal", "right": row.get("page_type") or ""},
        {"left": "group_title", "operation": "equal", "right": row.get("group_title") or ""},
        {"left": "section_title", "operation": "equal", "right": row.get("section_title") or ""},
        {"left": "placement_title", "operation": "equal", "right": row.get("placement_title") or ""},
        {"left": "checked_at", "operation": "like", "right": f"{checked_day(row.get('checked_at'))}%"},
    ])


def coze_upsert_rows(database_id, rows, key_fn, filter_fn):
    if not rows:
        return {"inserted": 0, "updated": 0}
    inserted = 0
    updated = 0
    grouped = {}
    for row in rows:
        grouped[key_fn(row)] = row
    for row in grouped.values():
        filter_payload = filter_fn(row)
        existing = coze_query_rows(database_id, page_num=1, page_size=20, filter_payload=filter_payload).get("items") or []
        if existing:
            primary_id = clean_text((existing[0] or {}).get("id") or "", 80)
            if primary_id:
                id_filter = row_filter([
                    {"left": "id", "operation": "equal", "right": primary_id},
                ])
                coze_update_rows(database_id, row, id_filter)
                updated += 1
                for duplicate in existing[1:]:
                    duplicate_id = clean_text((duplicate or {}).get("id") or "", 80)
                    if not duplicate_id:
                        continue
                    duplicate_filter = row_filter([
                        {"left": "id", "operation": "equal", "right": duplicate_id},
                    ])
                    coze_delete_rows(database_id, duplicate_filter)
                    time.sleep(0.05)
            else:
                coze_insert_rows(database_id, [row])
                updated += 1
        else:
            coze_insert_rows(database_id, [row])
            inserted += 1
    return {"inserted": inserted, "updated": updated}


def country_day_filter(country_code, checked_at):
    return row_filter([
        {"left": "country", "operation": "equal", "right": country_code or ""},
        {"left": "checked_at", "operation": "like", "right": f"{checked_day(checked_at)}%"},
    ])


def country_day_match_rows(country_code, checked_at):
    rows = []
    page_num = 1
    filter_payload = country_day_filter(country_code, checked_at)
    while page_num <= 30:
        queried = coze_query_rows(
            COZE_MATCHES_DB_ID,
            page_num=page_num,
            page_size=200,
            filter_payload=filter_payload,
            order_by=[{"field_name": "checked_at", "direction": "desc"}],
        )
        items = queried.get("items") or []
        rows.extend(items)
        if not queried.get("has_more"):
            break
        page_num += 1
    return rows


def country_day_run_rows(country_code, checked_at):
    rows = []
    page_num = 1
    filter_payload = country_day_filter(country_code, checked_at)
    while page_num <= 10:
        queried = coze_query_rows(
            COZE_RUNS_DB_ID,
            page_num=page_num,
            page_size=50,
            filter_payload=filter_payload,
            order_by=[{"field_name": "checked_at", "direction": "desc"}],
        )
        items = queried.get("items") or []
        rows.extend(items)
        if not queried.get("has_more"):
            break
        page_num += 1
    return rows


def build_run_row_from_match_rows(country_code, checked_at, rows, fallback_run_row=None):
    fallback_run_row = fallback_run_row or {}
    seen_games = {}
    page_pairs = set()
    latest_checked_at = checked_at or fallback_run_row.get("checked_at") or ""
    for row in rows:
        game_key = clean_text(row.get("game_key") or "", 160)
        game_name = clean_text(row.get("game_name") or "", 160)
        if game_key and game_name:
            seen_games[game_key] = game_name
        page_type = clean_text(row.get("page_type") or "", 32)
        if game_key and page_type:
            page_pairs.add((game_key, page_type))
        current_checked_at = row.get("checked_at") or ""
        if current_checked_at and current_checked_at > latest_checked_at:
            latest_checked_at = current_checked_at
    base_run_id = str(fallback_run_row.get("run_id") or "").split(":", 1)[0]
    return {
        "run_id": fallback_run_row.get("run_id") or f"{base_run_id}:{country_code}",
        "checked_at": latest_checked_at or checked_at,
        "country": country_code,
        "country_label": fallback_run_row.get("country_label") or "",
        "local_name": fallback_run_row.get("local_name") or "",
        "game_count": str(len(seen_games)),
        "page_count": str(len(page_pairs)),
        "match_count": str(len(rows)),
        "game_keys": "|".join(seen_games.keys()),
        "game_names": " / ".join(seen_games.values()),
        "status": "completed",
        "source": fallback_run_row.get("source") or "AppExpo",
    }


def coze_preserve_country_day_rows(run_rows, match_rows):
    if not run_rows and not match_rows:
        return {
            "runInsert": {"inserted": 0, "updated": 0},
            "matchInsert": {"inserted": 0, "updated": 0},
        }
    country_code = ""
    checked_at = ""
    if run_rows:
        country_code = run_rows[0].get("country") or ""
        checked_at = run_rows[0].get("checked_at") or ""
    elif match_rows:
        country_code = match_rows[0].get("country") or ""
        checked_at = match_rows[0].get("checked_at") or ""
    existing_match_rows = country_day_match_rows(country_code, checked_at)
    existing_match_map = {}
    existing_duplicate_rows = []
    for row in existing_match_rows:
        key = match_business_key(row)
        if not key:
            continue
        current = existing_match_map.get(key)
        if current is None:
            existing_match_map[key] = row
            continue
        preferred = choose_preferred_match_row(current, row)
        if row_checked_at(row) > row_checked_at(preferred):
            preferred = row
        if preferred is row:
            existing_duplicate_rows.append(current)
            existing_match_map[key] = row
        else:
            existing_duplicate_rows.append(row)

    desired_match_rows = {}
    for row in match_rows or []:
        desired_match_rows[match_business_key(row)] = row

    inserted = 0
    updated = 0
    skipped = 0
    merged_rows = []

    untouched_existing_rows = []
    for key, row in existing_match_map.items():
        if key not in desired_match_rows:
            untouched_existing_rows.append(row)

    for key, row in desired_match_rows.items():
        existing_row = existing_match_map.get(key) or {}
        existing_id = clean_text(existing_row.get("id") or "", 80)
        if existing_id:
            preferred = choose_preferred_match_row(existing_row, row)
            should_replace = preferred is row or row_checked_at(row) > row_checked_at(existing_row)
            if should_replace:
                coze_insert_rows(COZE_MATCHES_DB_ID, [row])
                delete_rows_by_ids(COZE_MATCHES_DB_ID, [existing_row])
                updated += 1
                merged_rows.append(row)
            else:
                skipped += 1
                merged_rows.append(existing_row)
        else:
            coze_insert_rows(COZE_MATCHES_DB_ID, [row])
            inserted += 1
            merged_rows.append(row)

    if existing_duplicate_rows:
        delete_rows_by_ids(COZE_MATCHES_DB_ID, existing_duplicate_rows)

    all_rows = untouched_existing_rows + merged_rows
    rebuilt_run_row = build_run_row_from_match_rows(
        country_code,
        checked_at,
        all_rows,
        dict(run_rows[0] if run_rows else {}),
    )
    existing_run_rows = country_day_run_rows(country_code, checked_at)
    existing_run_id = clean_text(((existing_run_rows[0] or {}).get("id") if existing_run_rows else "") or "", 80)
    if existing_run_id and inserted == 0:
        run_insert = {"inserted": 0, "updated": 0, "skipped": 1}
    elif existing_run_id:
        run_filter = row_filter([
            {"left": "id", "operation": "equal", "right": existing_run_id},
        ])
        coze_update_rows(COZE_RUNS_DB_ID, rebuilt_run_row, run_filter)
        run_insert = {"inserted": 0, "updated": 1, "skipped": 0}
    else:
        coze_insert_rows(COZE_RUNS_DB_ID, [rebuilt_run_row])
        run_insert = {"inserted": 1, "updated": 0, "skipped": 0}
    return {
        "runInsert": run_insert,
        "matchInsert": {"inserted": inserted, "updated": updated, "skipped": skipped},
    }


def local_run_id(sync_date, country_code):
    return f"{sync_date}:{country_code.upper()}"


def local_store_payload(payload):
    run_rows, match_rows = build_coze_rows(payload)
    inserted = 0
    updated = 0
    touched_pairs = set()
    now = utc_now_iso()
    with db_lock:
        connection = get_db_connection()
        try:
            for row in match_rows:
                row = dict(row)
                row["sync_date"] = checked_day(row.get("checked_at") or row.get("sync_date") or "")
                row["run_id"] = local_run_id(row["sync_date"], row.get("country") or "")
                business_key = match_business_key(row)
                if not business_key:
                    continue
                row["business_key"] = business_key
                touched_pairs.add((row["sync_date"], row.get("country") or ""))
                existing_rows = [
                    dict(item)
                    for item in connection.execute(
                        "SELECT * FROM placement_matches WHERE business_key = ?",
                        (business_key,),
                    ).fetchall()
                ]
                if not existing_rows:
                    existing_rows = [
                        dict(item)
                        for item in connection.execute(
	                            """
	                            SELECT * FROM placement_matches
	                            WHERE sync_date = ? AND country = ? AND page_type = ? AND app_id = ?
	                              AND media_mode = ? AND section_title = ? AND placement_title = ?
	                            """,
	                            (
	                                row.get("sync_date") or "",
	                                row.get("country") or "",
	                                row.get("page_type") or "",
	                                row.get("app_id") or "",
	                                row.get("media_mode") or "",
	                                row.get("section_title") or "",
	                                row.get("placement_title") or "",
	                            ),
	                        ).fetchall()
                    ]
                existing = None
                duplicate_existing = []
                for existing_row in existing_rows:
                    if existing is None:
                        existing = existing_row
                        continue
                    preferred = choose_preferred_match_row(existing, existing_row)
                    if preferred is existing_row:
                        duplicate_existing.append(existing)
                        existing = existing_row
                    else:
                        duplicate_existing.append(existing_row)
                if existing:
                    for duplicate in duplicate_existing:
                        connection.execute("DELETE FROM placement_matches WHERE id = ?", (duplicate.get("id"),))
                    connection.execute(
                        """
                        UPDATE placement_matches
                        SET business_key = ?, checked_at = ?, run_id = ?, game_key = ?, game_name = ?, app_id = ?,
                            country = ?, country_label = ?, local_name = ?, page_type = ?, page_label = ?,
                            group_title = ?, section_title = ?, placement_title = ?, subtitle = ?,
                            media_mode = ?, updated_at = ?, image = ?, app_icon = ?, raw_match = ?, touched_at = ?
                        WHERE id = ?
                        """,
                        (
                            business_key,
                            row.get("checked_at") or "",
                            row.get("run_id") or "",
                            row.get("game_key") or "",
                            row.get("game_name") or "",
                            row.get("app_id") or "",
                            row.get("country") or "",
                            row.get("country_label") or "",
                            row.get("local_name") or "",
                            row.get("page_type") or "",
                            row.get("page_label") or "",
                            row.get("group_title") or "",
                            row.get("section_title") or "",
                            row.get("placement_title") or "",
                            row.get("subtitle") or "",
                            row.get("media_mode") or "",
                            row.get("updated_at") or "",
                            row.get("image") or "",
                            row.get("app_icon") or "",
                            row.get("raw_match") or "",
                            now,
                            existing.get("id"),
                        ),
                    )
                    updated += 1
                else:
                    connection.execute(
                        """
                        INSERT INTO placement_matches (
                            business_key, sync_date, checked_at, run_id, game_key, game_name, app_id,
                            country, country_label, local_name, page_type, page_label, group_title,
                            section_title, placement_title, subtitle, media_mode, updated_at, image,
                            app_icon, raw_match, created_at, touched_at
                        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                        """,
                        (
                            business_key,
                            row.get("sync_date") or "",
                            row.get("checked_at") or "",
                            row.get("run_id") or "",
                            row.get("game_key") or "",
                            row.get("game_name") or "",
                            row.get("app_id") or "",
                            row.get("country") or "",
                            row.get("country_label") or "",
                            row.get("local_name") or "",
                            row.get("page_type") or "",
                            row.get("page_label") or "",
                            row.get("group_title") or "",
                            row.get("section_title") or "",
                            row.get("placement_title") or "",
                            row.get("subtitle") or "",
                            row.get("media_mode") or "",
                            row.get("updated_at") or "",
                            row.get("image") or "",
                            row.get("app_icon") or "",
                            row.get("raw_match") or "",
                            now,
                            now,
                        ),
                    )
                    inserted += 1

            for sync_date, country_code in touched_pairs:
                rows = [
                    dict(item)
                    for item in connection.execute(
                        """
                        SELECT game_key, game_name, page_type, checked_at, country_label, local_name
                        FROM placement_matches
                        WHERE sync_date = ? AND country = ?
                        ORDER BY checked_at DESC, id DESC
                        """,
                        (sync_date, country_code),
                    ).fetchall()
                ]
                if not rows:
                    continue
                rebuilt = build_run_row_from_match_rows(
                    country_code,
                    sync_date,
                    rows,
                    {
                        "run_id": local_run_id(sync_date, country_code),
                        "country_label": rows[0].get("country_label") or "",
                        "local_name": rows[0].get("local_name") or "",
                        "source": payload.get("source") or "AppExpo",
                    },
                )
                rebuilt["run_id"] = local_run_id(sync_date, country_code)
                rebuilt["sync_date"] = sync_date
                rebuilt["updated_at"] = now
                connection.execute(
                    """
                    INSERT INTO sync_runs (
                        run_id, sync_date, checked_at, country, country_label, local_name, game_count,
                        page_count, match_count, game_keys, game_names, status, source, updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(run_id) DO UPDATE SET
                        checked_at = excluded.checked_at,
                        country_label = excluded.country_label,
                        local_name = excluded.local_name,
                        game_count = excluded.game_count,
                        page_count = excluded.page_count,
                        match_count = excluded.match_count,
                        game_keys = excluded.game_keys,
                        game_names = excluded.game_names,
                        status = excluded.status,
                        source = excluded.source,
                        updated_at = excluded.updated_at
                    """,
                    (
                        rebuilt.get("run_id") or "",
                        rebuilt.get("sync_date") or "",
                        rebuilt.get("checked_at") or "",
                        rebuilt.get("country") or "",
                        rebuilt.get("country_label") or "",
                        rebuilt.get("local_name") or "",
                        int(rebuilt.get("game_count") or 0),
                        int(rebuilt.get("page_count") or 0),
                        int(rebuilt.get("match_count") or 0),
                        rebuilt.get("game_keys") or "",
                        rebuilt.get("game_names") or "",
                        rebuilt.get("status") or "completed",
                        rebuilt.get("source") or "AppExpo",
                        rebuilt.get("updated_at") or now,
                    ),
                )

            connection.commit()
        finally:
            connection.close()
    return {
        "target": "local-sqlite",
        "path": str(LOCAL_DB_FILE),
        "runRows": len(run_rows),
        "matchRows": len(match_rows),
        "runInsert": {"inserted": len(touched_pairs), "updated": 0, "skipped": 0},
        "matchInsert": {"inserted": inserted, "updated": updated, "skipped": 0},
    }


def catalog_app_ids_for_game(game_key):
    game = history_catalog_game_by_key(game_key) if game_key else None
    return {
        str(item or "").strip()
        for item in ((game or {}).get("appIds") or [])
        if str(item or "").strip()
    }


def sql_fetch_all(query, params=()):
    with db_lock:
        connection = get_db_connection()
        try:
            rows = connection.execute(query, params).fetchall()
            return [dict(row) for row in rows]
        finally:
            connection.close()


def build_history_match_where(game_key="", app_id="", country="", page_type="", date_from="", date_to=""):
    clauses = []
    params = []
    selected_app_ids = catalog_app_ids_for_game(game_key) if game_key else set()
    if game_key:
        if selected_app_ids:
            placeholders = ",".join("?" for _ in selected_app_ids)
            clauses.append(f"app_id IN ({placeholders})")
            params.extend(sorted(selected_app_ids))
        else:
            clauses.append("game_key = ?")
            params.append(game_key)
    if app_id:
        clauses.append("app_id = ?")
        params.append(app_id)
    if country:
        clauses.append("country = ?")
        params.append(country)
    if page_type:
        clauses.append("page_type = ?")
        params.append(page_type)
    if date_from:
        clauses.append("sync_date >= ?")
        params.append(date_from)
    if date_to:
        clauses.append("sync_date <= ?")
        params.append(date_to)
    where = f"WHERE {' AND '.join(clauses)}" if clauses else ""
    return where, params


def load_history_runs(page_num=1, page_size=20, keyword="", sync_date="", game_key=""):
    page_num = max(page_num, 1)
    page_size = min(max(page_size, 1), 100)
    items = sql_fetch_all(
        "SELECT * FROM sync_runs ORDER BY sync_date DESC, country ASC, checked_at DESC"
    )
    if sync_date:
        items = [item for item in items if clean_text(item.get("sync_date") or "", 20) == sync_date]
    if game_key:
        matched = load_history_matches(run_id="", game_key=game_key, page_num=1, page_size=1000, sync_date=sync_date).get("items") or []
        allowed_run_ids = {item.get("run_id") for item in matched if item.get("run_id")}
        items = [item for item in items if item.get("run_id") in allowed_run_ids]
    if keyword:
        normalized_keyword = normalize(keyword)
        items = [item for item in items if normalized_keyword in normalize(item.get("game_names") or "")]
    total_count = len(items)
    start = (page_num - 1) * page_size
    end = start + page_size
    items = items[start:end]
    return {
        "items": items,
        "totalCount": total_count,
        "hasMore": end < total_count,
        "pageNum": page_num,
        "pageSize": page_size,
    }


def load_history_games():
    latest_catalog = global_game_catalog().get("games") or []
    catalog_by_key = {str(game.get("key") or ""): game for game in latest_catalog}
    catalog_by_app_id = {}
    for game in latest_catalog:
        for app_id in (game.get("appIds") or [game.get("id")]):
            text = str(app_id or "").strip()
            if text and text not in catalog_by_app_id:
                catalog_by_app_id[text] = game
    catalog = {}
    items = sql_fetch_all(
        "SELECT game_key, game_name, app_id FROM placement_matches ORDER BY checked_at DESC, id DESC"
    )
    for item in items:
        game_key = clean_text(item.get("game_key") or "", 160)
        game_name = clean_text(item.get("game_name") or "", 160)
        app_id = clean_text(item.get("app_id") or "", 80)
        if not game_key and not app_id:
            continue
        matched_game = catalog_by_app_id.get(app_id) or catalog_by_key.get(game_key)
        canonical_key = str((matched_game or {}).get("key") or game_key or app_id)
        canonical_name = history_catalog_display_name(matched_game, game_name)
        existing = catalog.get(canonical_key) or {}
        app_ids = list(existing.get("appIds") or [])
        if app_id and app_id not in app_ids:
            app_ids.append(app_id)
        if matched_game:
            for extra_id in (matched_game.get("appIds") or []):
                extra_text = str(extra_id or "").strip()
                if extra_text and extra_text not in app_ids:
                    app_ids.append(extra_text)
        catalog[canonical_key] = {
            "key": canonical_key,
            "name": canonical_name or existing.get("name") or game_name or canonical_key,
            "appId": existing.get("appId") or app_id or "",
            "appIds": app_ids,
        }
    games = sorted(
        catalog.values(),
        key=lambda item: (
            history_game_priority_index(item.get("name") or ""),
            normalize(item.get("name") or ""),
        ),
    )
    return {
        "games": games,
        "totalCount": len(games),
    }


def load_history_run_games(run_id):
    latest_catalog = global_game_catalog().get("games") or []
    catalog_by_key = {str(game.get("key") or ""): game for game in latest_catalog}
    catalog_by_app_id = {}
    for game in latest_catalog:
        for app_id in (game.get("appIds") or [game.get("id")]):
            text = str(app_id or "").strip()
            if text and text not in catalog_by_app_id:
                catalog_by_app_id[text] = game
    seen = {}
    for item in sql_fetch_all(
        "SELECT run_id, game_key, game_name, app_id FROM placement_matches WHERE run_id = ? ORDER BY game_name ASC, id DESC",
        (run_id,),
    ):
        game_key = clean_text(item.get("game_key") or "", 160)
        game_name = clean_text(item.get("game_name") or "", 160)
        app_id = clean_text(item.get("app_id") or "", 80)
        if game_key and game_name:
            matched_game = catalog_by_app_id.get(app_id) or catalog_by_key.get(game_key)
            canonical_key = str((matched_game or {}).get("key") or game_key)
            canonical_name = history_catalog_display_name(matched_game, game_name)
            existing = seen.get(canonical_key) or {}
            app_ids = list(existing.get("appIds") or [])
            if app_id and app_id not in app_ids:
                app_ids.append(app_id)
            if matched_game:
                for extra_id in (matched_game.get("appIds") or []):
                    extra_text = str(extra_id or "").strip()
                    if extra_text and extra_text not in app_ids:
                        app_ids.append(extra_text)
            seen[canonical_key] = {
                "key": canonical_key,
                "name": canonical_name,
                "appId": existing.get("appId") or app_id or "",
                "appIds": app_ids,
            }
    games = sorted(
        seen.values(),
        key=lambda item: (
            history_game_priority_index(item.get("name") or ""),
            normalize(item.get("name") or ""),
        ),
    )
    return {"games": games, "totalCount": len(games)}


def parse_iso_day(value):
    text = str(value or "").strip()
    if not text:
        return ""
    return text[:10]


def history_catalog_game_by_key(game_key):
    target = str(game_key or "").strip()
    if not target:
        return None
    for game in global_game_catalog().get("games") or []:
        if str(game.get("key") or "") == target:
            return game
    return None


def history_catalog_display_name(game, fallback_name=""):
    if not game:
        return clean_text(fallback_name, 160)

    def is_latin_like(text):
        return bool(re.search(r"[A-Za-z]", text or ""))

    def looks_official(label):
        text = clean_text(label, 160)
        if not text:
            return False
        has_cjk_text = has_cjk(text)
        has_latin_text = bool(re.search(r"[A-Za-z]", text))
        has_joiner = " / " in text or " - " in text or "：" in text or ":" in text
        return has_joiner and has_cjk_text and has_latin_text

    chinese_name = clean_text(game.get("chineseName") or "", 160)
    if not chinese_name and has_cjk(fallback_name):
        chinese_name = clean_text(fallback_name, 160)

    official_dash = None
    slash_foreign = None
    candidates = []
    primary = clean_text(game.get("displayName") or game.get("chineseName") or fallback_name, 160)
    if primary:
        candidates.append(primary)

    for alias in game.get("aliases") or []:
        text = clean_text(alias, 160)
        if text:
            candidates.append(text)

    for entry in (game.get("countryEntries") or {}).values():
        text = clean_text((entry or {}).get("displayName") or (entry or {}).get("name") or "", 160)
        if text:
            candidates.append(text)

    for candidate in candidates:
        if looks_official(candidate) and " - " in candidate:
            official_dash = candidate
            break

    if official_dash:
        return official_dash

    if chinese_name:
        for candidate in candidates:
            text = clean_text(candidate, 160)
            if not text or not is_latin_like(text):
                continue
            if has_cjk(text):
                continue
            slash_foreign = text
            break

        if slash_foreign and normalize(slash_foreign) != normalize(chinese_name):
            return f"{chinese_name} / {slash_foreign}"

    for candidate in candidates:
        if looks_official(candidate):
            return candidate

    return primary or clean_text(fallback_name, 160)


def iterate_history_match_rows(game_key="", app_id="", country="", page_type=""):
    where, params = build_history_match_where(
        game_key=game_key,
        app_id=app_id,
        country=country,
        page_type=page_type,
    )
    for item in sql_fetch_all(
        f"SELECT * FROM placement_matches {where} ORDER BY checked_at DESC, id DESC",
        tuple(params),
    ):
        yield item


def match_in_date_range(item, date_from="", date_to=""):
    current = parse_iso_day(item.get("checked_at") or "")
    if not current:
        return False
    if date_from and current < date_from:
        return False
    if date_to and current > date_to:
        return False
    return True


def load_history_countries(game_key="", app_id="", date_from="", date_to="", page_type=""):
    countries = {}
    for item in iterate_history_match_rows(game_key=game_key, app_id=app_id, page_type=page_type):
        if not match_in_date_range(item, date_from=date_from, date_to=date_to):
            continue
        code = clean_text(item.get("country") or "", 20)
        if not code:
            continue
        bucket = countries.setdefault(code, {
            "code": code,
            "label": item.get("country_label") or "",
            "localName": item.get("local_name") or "",
            "matchCount": 0,
            "games": set(),
        })
        bucket["matchCount"] += 1
        if item.get("game_name"):
            bucket["games"].add(item.get("game_name"))
    rows = []
    for bucket in countries.values():
        rows.append({
            "code": bucket["code"],
            "label": bucket["label"],
            "localName": bucket["localName"],
            "matchCount": bucket["matchCount"],
            "gameCount": len(bucket["games"]),
        })
    rows.sort(key=lambda item: (-item["matchCount"], item["code"]))
    return {"countries": rows, "totalCount": len(rows)}


def load_history_analytics(game_key="", app_id="", country="", page_type="", date_from="", date_to=""):
    items = []
    for item in iterate_history_match_rows(game_key=game_key, app_id=app_id, country=country, page_type=page_type):
        if not match_in_date_range(item, date_from=date_from, date_to=date_to):
            continue
        items.append(item)
    groups = {}
    for item in items:
        key = item.get("country") or ""
        bucket = groups.setdefault(key, {
            "country": item.get("country") or "",
            "countryLabel": item.get("country_label") or "",
            "localName": item.get("local_name") or "",
            "matches": [],
        })
        bucket["matches"].append(item)
    grouped = sorted(groups.values(), key=lambda item: (-len(item["matches"]), item["country"]))
    summary = {
        "matchCount": len(items),
        "countryCount": len(grouped),
        "todayCount": sum(1 for item in items if item.get("page_type") == "today"),
        "gamesCount": sum(1 for item in items if item.get("page_type") == "games"),
    }
    return {
        "summary": summary,
        "groups": grouped,
        "items": items,
    }


def load_history_analytics_batch(selections=None, date_from="", date_to="", page_type=""):
    selections = selections if isinstance(selections, list) else []
    clauses = []
    params = []
    seen = set()
    for selection in selections[:500]:
        if not isinstance(selection, dict):
            continue
        country = clean_text(selection.get("country") or selection.get("countryCode") or "", 20).upper()
        app_id = clean_text(selection.get("appId") or selection.get("app_id") or "", 80)
        game_key = clean_text(selection.get("gameKey") or selection.get("game_key") or "", 160)
        if country and app_id:
            key = ("country_app", country, app_id)
            if key in seen:
                continue
            seen.add(key)
            clauses.append("(country = ? AND app_id = ?)")
            params.extend([country, app_id])
        elif country and game_key:
            key = ("country_game", country, game_key)
            if key in seen:
                continue
            seen.add(key)
            clauses.append("(country = ? AND game_key = ?)")
            params.extend([country, game_key])
        elif app_id:
            key = ("app", app_id)
            if key in seen:
                continue
            seen.add(key)
            clauses.append("app_id = ?")
            params.append(app_id)
        elif game_key:
            key = ("game", game_key)
            if key in seen:
                continue
            seen.add(key)
            selected_app_ids = catalog_app_ids_for_game(game_key)
            if selected_app_ids:
                placeholders = ",".join("?" for _ in selected_app_ids)
                clauses.append(f"app_id IN ({placeholders})")
                params.extend(sorted(selected_app_ids))
            else:
                clauses.append("game_key = ?")
                params.append(game_key)
    if not clauses:
        return {
            "summary": {
                "matchCount": 0,
                "countryCount": 0,
                "todayCount": 0,
                "gamesCount": 0,
                "dateCount": 0,
                "dates": [],
            },
            "groups": [],
            "items": [],
        }
    where_clauses = [f"({' OR '.join(clauses)})"]
    where_params = list(params)
    if page_type:
        where_clauses.append("page_type = ?")
        where_params.append(page_type)
    if date_from:
        where_clauses.append("sync_date >= ?")
        where_params.append(date_from)
    if date_to:
        where_clauses.append("sync_date <= ?")
        where_params.append(date_to)
    where = f"WHERE {' AND '.join(where_clauses)}"
    items = sql_fetch_all(
        f"SELECT * FROM placement_matches {where} ORDER BY checked_at DESC, id DESC",
        tuple(where_params),
    )
    groups = {}
    dates = set()
    for item in items:
        sync_date = parse_iso_day(item.get("sync_date") or item.get("checked_at") or "")
        if sync_date:
            dates.add(sync_date)
        key = item.get("country") or ""
        bucket = groups.setdefault(key, {
            "country": item.get("country") or "",
            "countryLabel": item.get("country_label") or "",
            "localName": item.get("local_name") or "",
            "matches": [],
        })
        bucket["matches"].append(item)
    grouped = sorted(groups.values(), key=lambda item: (-len(item["matches"]), item["country"]))
    ordered_dates = sorted(dates, reverse=True)
    summary = {
        "matchCount": len(items),
        "countryCount": len(grouped),
        "todayCount": sum(1 for item in items if item.get("page_type") == "today"),
        "gamesCount": sum(1 for item in items if item.get("page_type") == "games"),
        "dateCount": len(ordered_dates),
        "dates": ordered_dates,
    }
    return {
        "summary": summary,
        "groups": grouped,
        "items": items,
    }


def load_history_matches(run_id="", game_key="", page_num=1, page_size=100, country="", page_type="", sync_date=""):
    page_num = max(page_num, 1)
    page_size = min(max(page_size, 1), 200)
    date_from = sync_date or ""
    date_to = sync_date or ""
    where, params = build_history_match_where(
        game_key=game_key,
        app_id="",
        country=country,
        page_type=page_type,
        date_from=date_from,
        date_to=date_to,
    )
    if run_id:
        if where:
            where += " AND run_id = ?"
        else:
            where = "WHERE run_id = ?"
        params.append(run_id)
    rows = sql_fetch_all(
        f"SELECT * FROM placement_matches {where} ORDER BY checked_at DESC, id DESC",
        tuple(params),
    )
    total_count = len(rows)
    start = (page_num - 1) * page_size
    end = start + page_size
    items = rows[start:end]
    return {
        "items": items,
        "totalCount": total_count,
        "hasMore": end < total_count,
        "pageNum": page_num,
        "pageSize": page_size,
    }


def normalize_match_for_sync(match):
    return {
        "country": match.get("country") or "",
        "countryLabel": match.get("countryLabel") or "",
        "localName": match.get("localName") or "",
        "pageType": match.get("pageType") or "",
        "pageLabel": match.get("pageLabel") or "",
        "sectionTitle": match.get("sectionTitle") or "",
        "sectionSubtitle": match.get("sectionSubtitle") or "",
        "groupTitle": match.get("groupTitle") or "",
        "groupSubtitle": match.get("groupSubtitle") or "",
        "headerTitle": match.get("headerTitle") or "",
        "placementType": match.get("placementType") or "",
        "placementTitle": match.get("placementTitle") or "",
        "subtitle": match.get("subtitle") or "",
        "description": match.get("description") or "",
        "eventStatus": match.get("eventStatus") or "",
        "eventBadgeKind": match.get("eventBadgeKind") or "",
        "eventKind": match.get("eventKind") or "",
        "eventStartDate": match.get("eventStartDate") or "",
        "eventEndDate": match.get("eventEndDate") or "",
        "eventRequirement": match.get("eventRequirement") or "",
        "callToAction": match.get("callToAction") or "",
        "appId": match.get("id") or "",
        "appTitle": match.get("appTitle") or "",
        "appSubtitle": match.get("appSubtitle") or "",
        "appIcon": match.get("appIcon") or match.get("iconImage") or "",
        "image": match.get("image") or "",
        "heroImage": match.get("heroImage") or "",
        "heroRibbon": match.get("heroRibbon") or "",
        "heroEyebrow": match.get("heroEyebrow") or "",
        "heroTitle": match.get("heroTitle") or "",
        "heroDescription": match.get("heroDescription") or "",
        "buttonNote": match.get("buttonNote") or "",
        "mediaMode": match.get("mediaMode") or "",
        "position": match.get("position"),
        "modulePosition": match.get("modulePosition"),
        "itemPosition": match.get("itemPosition"),
        "groupItemCount": match.get("groupItemCount"),
        "contentPosition": match.get("contentPosition"),
        "overallPosition": match.get("overallPosition"),
        "updatedAt": match.get("updatedAt") or "",
        "checkedAt": match.get("checkedAt") or "",
        "editorialKind": match.get("editorialKind") or "",
        "path": match.get("path") or "",
        "textSnippet": match.get("textSnippet") or "",
    }


def has_placement_image(match):
    return bool((match or {}).get("image"))


def normalize_page_result_for_sync(result):
    matches = [match for match in (result.get("matches") or []) if has_placement_image(match)]
    return {
        "country": result.get("country") or "",
        "countryLabel": result.get("countryLabel") or "",
        "localName": result.get("localName") or "",
        "pageType": result.get("pageType") or "",
        "pageLabel": result.get("pageLabel") or "",
        "url": result.get("url") or "",
        "found": bool(matches),
        "error": result.get("error") or "",
        "matchCount": len(matches),
        "matches": [normalize_match_for_sync(match) for match in matches],
    }


def build_sync_payload(body):
    results = body.get("results") or []
    games = body.get("games") or []
    checked_at = body.get("checkedAt") or datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
    run_id = body.get("runId") or f"appexpo-{int(time.time())}"

    if games:
        normalized_games = []
        for game in games:
            normalized_games.append({
                "gameKey": game.get("gameKey") or game.get("key") or "",
                "gameName": game.get("gameName") or game.get("displayName") or "",
                "appId": str(game.get("appId") or game.get("id") or ""),
                "aliases": [clean_text(item, 120) for item in (game.get("aliases") or []) if clean_text(item, 120)],
                "appIcon": str(game.get("appIcon") or game.get("icon") or ""),
                "checkedAt": game.get("checkedAt") or checked_at,
                "results": [normalize_page_result_for_sync(result) for result in (game.get("results") or [])],
            })
        return {
            "schemaVersion": "1.0",
            "source": "AppExpo",
            "mode": body.get("mode") or "scheduled-full-sync",
            "runId": run_id,
            "checkedAt": checked_at,
            "games": normalized_games,
        }

    return {
        "schemaVersion": "1.0",
        "source": "AppExpo",
        "mode": body.get("mode") or ("cross-country" if body.get("crossCountry") else "single-game"),
        "runId": run_id,
        "checkedAt": checked_at,
        "game": {
            "gameKey": body.get("gameKey") or "",
            "gameName": body.get("gameName") or "",
            "appId": str(body.get("appId") or ""),
        },
        "results": [normalize_page_result_for_sync(result) for result in results],
    }


def build_country_sync_payloads(run_id, checked_at, catalog, results_by_game):
    country_payloads = []
    countries = sorted({country_code for game in catalog for country_code in game["countryCodes"]})
    for country_code in countries:
        games_payload = []
        for game in catalog:
            if country_code not in game["countryCodes"]:
                continue
            game_results = [
                result for result in (results_by_game.get(game["key"]) or [])
                if result.get("country") == country_code and (result.get("matches") or [])
            ]
            if not game_results:
                continue
            country_entry = (game.get("countryEntries") or {}).get(country_code) or {}
            games_payload.append({
                "gameKey": game["key"],
                "gameName": country_entry.get("displayName") or country_entry.get("name") or game["displayName"],
                "appId": str(country_entry.get("appId") or game["id"]),
                "aliases": country_entry.get("aliases") or game.get("aliases") or [],
                "appIcon": str(country_entry.get("icon") or ""),
                "checkedAt": checked_at,
                "results": game_results,
            })
        if not games_payload:
            continue
        country_payloads.append(build_sync_payload({
            "mode": "scheduled-full-sync",
            "runId": run_id,
            "checkedAt": checked_at,
            "games": games_payload,
        }))
    return country_payloads


def build_coze_rows(payload):
    games = payload.get("games") or []
    if not games and payload.get("game"):
        games = [{
            "gameKey": (payload.get("game") or {}).get("gameKey") or "",
            "gameName": (payload.get("game") or {}).get("gameName") or "",
            "appId": (payload.get("game") or {}).get("appId") or "",
            "checkedAt": payload.get("checkedAt") or "",
            "results": payload.get("results") or [],
        }]
    run_rows = []
    match_rows = []
    deduped_match_rows = {}
    country_groups = {}
    for game in games:
        for result in (game.get("results") or []):
            matches = result.get("matches") or []
            if not matches:
                continue
            country_code = result.get("country") or ""
            if not country_code:
                continue
            bucket = country_groups.setdefault(country_code, {
                "checked_at": game.get("checkedAt") or payload.get("checkedAt") or "",
                "country": country_code,
                "country_label": result.get("countryLabel") or "",
                "local_name": result.get("localName") or "",
                "games": {},
                "page_count": 0,
            })
            bucket["page_count"] += 1
            bucket["games"][game.get("gameKey") or ""] = game.get("gameName") or ""
            for match in matches:
                if not has_placement_image(match):
                    continue
                raw_match = dict(match)
                raw_match["countryLabel"] = result.get("countryLabel") or ""
                raw_match["localName"] = result.get("localName") or ""
                row = {
                    "sync_date": checked_day(match.get("checkedAt") or payload.get("checkedAt") or ""),
                    "checked_at": match.get("checkedAt") or payload.get("checkedAt") or "",
                    "run_id": f"{payload.get('runId') or ''}:{country_code}",
                    "game_key": game.get("gameKey") or "",
                    "game_name": game.get("gameName") or "",
                    "app_id": match.get("appId") or game.get("appId") or "",
                    "country": result.get("country") or match.get("country") or "",
                    "country_label": result.get("countryLabel") or "",
                    "local_name": result.get("localName") or "",
                    "page_type": result.get("pageType") or match.get("pageType") or "",
                    "page_label": result.get("pageLabel") or match.get("pageLabel") or "",
                    "group_title": match.get("groupTitle") or "",
                    "section_title": match.get("sectionTitle") or "",
                    "group_title": match.get("groupTitle") or "",
                    "placement_title": match.get("placementTitle") or "",
                    "subtitle": match.get("subtitle") or "",
                    "media_mode": match.get("mediaMode") or "",
                    "updated_at": match.get("updatedAt") or "",
                    "image": match.get("image") or match.get("heroImage") or "",
                    "app_icon": match.get("appIcon") or match.get("iconImage") or "",
                    "raw_match": json.dumps(raw_match, ensure_ascii=False),
                }
                dedupe_key = match_business_key(row)
                existing = deduped_match_rows.get(dedupe_key)
                if existing is None:
                    deduped_match_rows[dedupe_key] = row
                else:
                    deduped_match_rows[dedupe_key] = choose_preferred_match_row(existing, row)
    match_rows = list(deduped_match_rows.values())
    country_match_counts = {}
    for row in match_rows:
        country_code = row.get("country") or ""
        if not country_code:
            continue
        country_match_counts[country_code] = country_match_counts.get(country_code, 0) + 1
    for country_code, bucket in country_groups.items():
        game_pairs = [(key, name) for key, name in bucket["games"].items() if key and name]
        run_rows.append({
            "run_id": f"{payload.get('runId') or ''}:{country_code}",
            "checked_at": bucket["checked_at"],
            "country": country_code,
            "country_label": bucket["country_label"],
            "local_name": bucket["local_name"],
            "game_count": str(len(game_pairs)),
            "page_count": str(bucket["page_count"]),
            "match_count": str(country_match_counts.get(country_code, 0)),
            "game_keys": "|".join(key for key, _name in game_pairs),
            "game_names": " / ".join(name for _key, name in game_pairs),
            "status": "completed",
            "source": payload.get("source") or "AppExpo",
        })
    return run_rows, match_rows


def post_sync_payload(payload):
    return local_store_payload(payload)


def apple_headers(country):
    return {
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Safari/605.1.15",
        "accept": "application/json,text/plain,*/*",
        "accept-language": f"{country['locale']},zh-CN;q=0.9,en;q=0.8",
        "cache-control": "no-cache, max-age=0",
        "pragma": "no-cache",
        "origin": "https://apps.apple.com",
        "referer": f"https://apps.apple.com/{country['path']}/iphone/games?l={urllib.parse.quote(country['locale'])}",
        "x-apple-store-front": country["storefront"],
    }


def retry_delay(attempt, error):
    if isinstance(error, AppleCapacityError) or re.search(r"429|capacity exceeded", str(error), re.I):
        return 1.4 + attempt * 2.2
    return 0.65 + attempt * 0.65


def fetch_text(url, headers, timeout=30):
    request = urllib.request.Request(url, headers=headers, method="GET")
    with urllib.request.urlopen(request, timeout=timeout) as response:
        return response.status, response.read(), response.headers


def fetch_json(url, country, retries=2, editorial=False):
    last_error = None
    headers = apple_headers(country)
    if editorial:
        headers = {
            "user-agent": headers["user-agent"],
            "accept": "application/json,text/plain,*/*",
            "accept-language": f"{country['locale']},zh-CN;q=0.9,en;q=0.8",
            "cookie": "geo=SG",
            "cache-control": "no-cache, max-age=0",
            "pragma": "no-cache",
        }
    for attempt in range(retries + 1):
        try:
            status, raw, _resp_headers = fetch_text(url, headers)
            text = raw.decode("utf-8", "replace")
            if re.match(r"^API capacity exceeded", text, re.I) or status == 429:
                raise AppleCapacityError(f"HTTP 429: {text[:180]}")
            if status >= 400:
                raise RuntimeError(f"HTTP {status}: {text[:180]}")
            data = json.loads(text)
            if re.search(r"capacity exceeded", json.dumps(data, ensure_ascii=False)[:600], re.I):
                raise AppleCapacityError("Apple API capacity exceeded")
            return data
        except urllib.error.HTTPError as error:
            body = error.read().decode("utf-8", "replace")
            last_error = AppleCapacityError(f"HTTP {error.code}: {body[:180]}") if error.code == 429 else RuntimeError(f"HTTP {error.code}: {body[:180]}")
        except (urllib.error.URLError, json.JSONDecodeError, RuntimeError, AppleCapacityError) as error:
            last_error = error
        if attempt < retries:
            time.sleep(retry_delay(attempt, last_error))
    raise last_error


def lookup_app_icon(app_id, country_code):
    app_id = clean_text(app_id, 40)
    country_code = clean_text(country_code, 8).upper()
    if not app_id or country_code not in COUNTRIES:
        return ""
    cache_key = f"{country_code}:{app_id}"
    if cache_key in app_icon_lookup_cache:
        return app_icon_lookup_cache[cache_key]
    country = COUNTRIES[country_code]
    lookup_url = f"https://itunes.apple.com/{country['path']}/lookup?{urllib.parse.urlencode({'id': app_id, 'entity': 'software'})}"
    icon = ""
    try:
        status, raw, _resp_headers = fetch_text(lookup_url, apple_headers(country), timeout=20)
        if status < 400:
            data = json.loads(raw.decode("utf-8", "replace"))
            result = (data.get("results") or [{}])[0]
            icon = result.get("artworkUrl512") or result.get("artworkUrl100") or result.get("artworkUrl60") or ""
    except Exception:
        icon = ""
    app_icon_lookup_cache[cache_key] = icon
    return icon


def fetch_image_data_url(image_url):
    parsed = urllib.parse.urlparse(image_url)
    if parsed.scheme != "https":
        raise RuntimeError("仅支持 HTTPS 图片")
    request = urllib.request.Request(image_url, headers={
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/18.0 Safari/605.1.15",
        "accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
    })
    with urllib.request.urlopen(request, timeout=20) as response:
        content_type = response.headers.get("content-type") or "image/jpeg"
        if not content_type.startswith("image/"):
            raise RuntimeError("图片类型不正确")
        raw = response.read(8 * 1024 * 1024)
    encoded = base64.b64encode(raw).decode("ascii")
    return f"data:{content_type};base64,{encoded}"


def run_python_export(payload):
    with tempfile.TemporaryDirectory(prefix="appexpo-export-") as temp_dir:
        input_path = Path(temp_dir) / "input.json"
        output_path = Path(temp_dir) / "export.png"
        input_path.write_text(json.dumps(payload, ensure_ascii=False), encoding="utf-8")
        completed = subprocess.run(
            ["python3", str(SCRIPTS_DIR / "export_result.py"), str(input_path), str(output_path)],
            capture_output=True,
            text=True,
            timeout=60,
            check=False,
        )
        if completed.returncode != 0:
            raise RuntimeError((completed.stderr or completed.stdout or "导出图片失败").strip())
        return output_path.read_bytes()


def editorial_url(country_code, page_type):
    country = COUNTRIES[country_code]
    page = PAGE_TYPES[page_type]
    url = urllib.parse.urlparse(f"https://apps.apple.com/api/apps/v1/editorial/{country['path']}/{page['path']}")
    params = urllib.parse.parse_qsl(page["query"], keep_blank_values=True)
    params.append(("l", country["locale"]))
    if country.get("editorialPlatforms"):
        params = [(k, v) for k, v in params if k != "additionalPlatforms"]
        params.append(("additionalPlatforms", country["editorialPlatforms"]))
    return urllib.parse.urlunparse(url._replace(query=urllib.parse.urlencode(params)))


def is_artwork_url(value):
    text = str(value or "")
    return "{w}" in text or "mzstatic.com/image/" in text or bool(re.search(r"\.(png|jpe?g|webp)(\?|$)", text, re.I))


def artwork_dimensions(value, size):
    width = value.get("width") if isinstance(value, dict) else None
    height = value.get("height") if isinstance(value, dict) else None
    if not isinstance(width, (int, float)) or not isinstance(height, (int, float)) or width <= 0 or height <= 0:
        return size, size
    if width >= height:
        return size, max(1, round(size * height / width))
    return max(1, round(size * width / height)), size


def fill_artwork_template(url, width, height=None):
    height = height or width
    return re.sub(r"\{[^}]+\}", "", str(url).replace("{w}", str(width)).replace("{h}", str(height)).replace("{f}", "jpg").replace("{c}", "bb"))


def resolve_artwork_url(value, size=720):
    if not value:
        return ""
    if isinstance(value, str):
        return fill_artwork_template(value, size, size) if is_artwork_url(value) else ""
    if isinstance(value, list):
        for item in value:
            found = resolve_artwork_url(item, size)
            if found:
                return found
        return ""
    if not isinstance(value, dict):
        return ""
    direct = value.get("urlTemplate") or value.get("src") or value.get("url") or value.get("href")
    if direct and is_artwork_url(direct):
        width, height = artwork_dimensions(value, size)
        return fill_artwork_template(direct, width, height)
    for candidate_key in ("dictionary", "artwork", "editorialArtwork", "iconArtwork", "customArtwork", "productArtwork", "backgroundArtwork", "lockupArtwork", "image", "platformAttributes", "ios", "iphone", "ipad"):
        found = resolve_artwork_url(value.get(candidate_key), size)
        if found:
            return found
    return ""


def artwork_from_keys(value, keys, size=960):
    if not isinstance(value, dict):
        return ""
    for key in keys:
        current = value
        for part in key.split("."):
            current = current.get(part) if isinstance(current, dict) else None
        found = resolve_artwork_url(current, size)
        if found:
            return found
    return ""


def editorial_image_from_node(node):
    artwork = ((node or {}).get("attributes") or {}).get("editorialArtwork") or ((node or {}).get("attributes") or {}).get("artwork") or (node or {}).get("editorialArtwork")
    return artwork_from_keys(artwork, [
        "dayCard", "generalCard", "storyCenteredStatic16x9", "universalAStatic16x9",
        "bannerUber", "subscriptionHero", "storeFlowcase", "categoryDetailStatic16x9",
        "searchCategoryBrick", "contentGraphicTrimmed", "productPageHero"
    ], 2160) or resolve_artwork_url(artwork, 2160)


def event_image_from_node(node):
    attrs = (node or {}).get("attributes") or {}
    return artwork_from_keys(attrs, [
        "lockupArtwork",
        "productArtwork",
        "lockupVideo.previewFrame",
        "productVideo.previewFrame",
    ], 1180)


def event_video_from_node(node):
    attrs = (node or {}).get("attributes") or {}
    for key in ("lockupVideo", "productVideo"):
        video = attrs.get(key)
        if isinstance(video, dict) and video.get("video"):
            return video.get("video") or ""
    return ""


def app_icon_from_node(node, fallback=""):
    attrs = (node or {}).get("attributes") if isinstance(node, dict) else node or {}
    placeholder = ""
    for key in [
        "platformAttributes.ios.artwork",
        "platformAttributes.ios.customAttributes.default.default.customArtwork",
        "platformAttributes.iphone.artwork",
        "artwork",
        "customArtwork",
        "platformAttributes.ios.iconArtwork",
        "iconArtwork",
        "customIconArtwork",
    ]:
        found = artwork_from_keys(attrs, [key], 512)
        if not found:
            continue
        if "placeholder" in found.lower():
            placeholder = placeholder or found
            continue
        return found
    return fallback or placeholder


def field_value(node, keys):
    if not isinstance(node, dict):
        return ""
    for key in keys:
        current = node
        for part in key.split("."):
            current = current.get(part) if isinstance(current, dict) else None
        if isinstance(current, str) and current.strip():
            return clean_text(current, 260)
    return ""


def title_from_node(node):
    return field_value(node, [
        "attributes.name", "attributes.title", "attributes.headerName", "attributes.displayName",
        "attributes.editorialNotes.name", "attributes.enrichedEditorialNotes.name", "attributes.artistName",
        "attributes.offerName", "attributes.label", "name", "title", "headerName", "displayName"
    ])


def subtitle_from_node(node):
    return field_value(node, [
        "attributes.subtitle", "attributes.editorialNotes.tagline", "attributes.tagline", "attributes.headerTagline",
        "attributes.platformAttributes.ios.subtitle", "attributes.platformAttributes.iphone.subtitle",
        "attributes.editorialNotes.short", "attributes.enrichedEditorialNotes.short",
        "attributes.editorialNotes.badge", "attributes.enrichedEditorialNotes.badge",
        "attributes.shortEditorialNotes.standard", "attributes.shortEditorialNotes",
        "attributes.editorialNotes.standard", "attributes.description.standard",
        "attributes.description", "subtitle", "tagline", "description"
    ])


def source_from_node(node):
    return {
        "title": title_from_node(node),
        "subtitle": subtitle_from_node(node),
        "updatedAt": (((node or {}).get("attributes") or {}).get("lastModifiedDate") or ""),
        "displayStyle": (((node or {}).get("attributes") or {}).get("displayStyle") or ((node or {}).get("attributes") or {}).get("cardDisplayStyle") or ""),
        "kind": (((node or {}).get("attributes") or {}).get("editorialElementKind") or ((node or {}).get("attributes") or {}).get("kind") or (node or {}).get("type") or ""),
    }


def presentation_from_editorial_item(node):
    attrs = (node or {}).get("attributes") or {}
    notes = attrs.get("editorialNotes") or attrs.get("enrichedEditorialNotes") or {}
    return {
        "title": clean_text(notes.get("name") or attrs.get("headerName") or attrs.get("name") or attrs.get("title") or ""),
        "subtitle": clean_text(notes.get("tagline") or notes.get("short") or attrs.get("headerTagline") or attrs.get("tagline") or attrs.get("subtitle") or ""),
        "badge": clean_text(notes.get("badge") or attrs.get("label") or ""),
        "callToAction": clean_text(notes.get("callToAction") or ""),
    }


def presentation_from_editorial_group(node, fallback_title=""):
    attrs = (node or {}).get("attributes") or {}
    notes = attrs.get("editorialNotes") or attrs.get("enrichedEditorialNotes") or {}
    return {
        "title": clean_text(notes.get("name") or attrs.get("name") or attrs.get("title") or fallback_title or ""),
        "subtitle": clean_text(notes.get("tagline") or notes.get("short") or attrs.get("headerTagline") or attrs.get("tagline") or ""),
        "badge": clean_text(notes.get("badge") or attrs.get("label") or ""),
    }


def group_title_from_section(node, fallback_title=""):
    attrs = (node or {}).get("attributes") or {}
    notes = attrs.get("editorialNotes") or attrs.get("enrichedEditorialNotes") or {}
    return clean_text(notes.get("name") or attrs.get("name") or attrs.get("title") or fallback_title or "")


def group_subtitle_from_section(node):
    attrs = (node or {}).get("attributes") or {}
    notes = attrs.get("editorialNotes") or attrs.get("enrichedEditorialNotes") or {}
    return clean_text(notes.get("tagline") or notes.get("short") or attrs.get("headerTagline") or attrs.get("tagline") or attrs.get("subtitle") or "")


def event_status(attributes):
    now = time.time() * 1000
    start_raw = attributes.get("promotionStartDate") or attributes.get("startDate") or ""
    end_raw = attributes.get("endDate") or ""
    start = parse_timestamp(start_raw)
    end = parse_timestamp(end_raw)
    if end is not None and now > end:
        return "已结束"
    if start is not None and now < start:
        return "即将开始"
    return "进行中"


def parse_timestamp(value):
    try:
        dt = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        return dt.timestamp() * 1000
    except Exception:
        return None


def event_details_from_node(event_node, app_node):
    if not event_node:
        return None
    attrs = event_node.get("attributes") or {}
    return {
        "id": event_node.get("id") or "",
        "title": attrs.get("name") or "活动",
        "subtitle": attrs.get("subtitle") or "",
        "description": (attrs.get("description") or {}).get("standard") if isinstance(attrs.get("description"), dict) else (attrs.get("description") or ""),
        "kind": attrs.get("kind") or attrs.get("eventKind") or "活动",
        "badgeKind": attrs.get("badgeKind") or "",
        "status": event_status(attrs),
        "startDate": attrs.get("startDate") or attrs.get("promotionStartDate") or "",
        "endDate": attrs.get("endDate") or "",
        "requirement": attrs.get("requirement") or "",
        "image": event_image_from_node(event_node),
        "video": event_video_from_node(event_node),
        "appTitle": title_from_node(app_node),
        "appSubtitle": subtitle_from_node(app_node),
        "appIcon": app_icon_from_node(app_node),
        "appId": (app_node or {}).get("id") or "",
        "url": attrs.get("url") or "",
    }


def node_text_blob(node):
    pieces = []
    keys = {"name", "title", "headerName", "displayName", "subtitle", "tagline", "headerTagline", "description", "standard", "short"}

    def visit(value, depth=0):
        if depth > 4 or value is None:
            return
        if isinstance(value, str):
            if len(value) <= 320:
                pieces.append(value)
            return
        if isinstance(value, list):
            for item in value[:12]:
                visit(item, depth + 1)
            return
        if isinstance(value, dict):
            for key, next_value in value.items():
                if key in keys or "Name" in key or "Title" in key or "Notes" in key:
                    visit(next_value, depth + 1)

    visit(node)
    return " · ".join(pieces)


def today_recommendations(content):
    return ((((content or {}).get("meta") or {}).get("associations") or {}).get("recommendations") or {}).get("data") or \
        ((((content or {}).get("relationships") or {}).get("recommendations") or {}).get("data") or []) or \
        ((((content or {}).get("recommendations") or {}).get("data")) or [])


def direct_placement_parts(item):
    parts = []
    if not isinstance(item, dict):
        return parts
    if item.get("type") == "apps":
        return [{"appNode": item, "eventNode": None, "item": item, "relationshipKey": "", "relatedIndex": 0}]
    if item.get("type") == "app-events":
        app_node = next((child for child in ((((item.get("relationships") or {}).get("app") or {}).get("data")) or []) if child.get("type") == "apps"), None)
        return [{"appNode": app_node, "eventNode": item, "item": item, "relationshipKey": "", "relatedIndex": 0}] if app_node else []
    for key in ("card-contents", "primary-content", "marketing-items"):
        related = ((((item.get("relationships") or {}).get(key) or {}).get("data")) or [])
        for related_index, related_item in enumerate(related):
            if related_item.get("type") == "apps":
                parts.append({"appNode": related_item, "eventNode": None, "item": item, "relationshipKey": key, "relatedIndex": related_index})
            elif related_item.get("type") == "app-events":
                app_node = next((child for child in ((((related_item.get("relationships") or {}).get("app") or {}).get("data")) or []) if child.get("type") == "apps"), None)
                if app_node:
                    parts.append({"appNode": app_node, "eventNode": related_item, "item": item, "relationshipKey": key, "relatedIndex": related_index})
    return parts


def app_name_text_blob(app_node):
    if not isinstance(app_node, dict) or app_node.get("type") != "apps":
        return ""
    attrs = app_node.get("attributes") or {}
    return normalize("".join(filter(None, [
        attrs.get("name"),
        ((((attrs.get("platformAttributes") or {}).get("ios") or {}).get("name")) or ""),
        ((((attrs.get("platformAttributes") or {}).get("iphone") or {}).get("name")) or ""),
        ((((attrs.get("platformAttributes") or {}).get("ipad") or {}).get("name")) or ""),
        app_node.get("name") or "",
    ])))


def games_section_title(section, item, app_title, media_mode):
    raw_title = title_from_node(section)
    if raw_title:
        return raw_title
    presentation = presentation_from_editorial_item(item)
    section_attrs = (section or {}).get("attributes") or {}
    section_kind = str(section_attrs.get("editorialElementKind") or "")
    item_display = str(((item or {}).get("attributes") or {}).get("displayStyle") or "")

    # CN Games 里像“精选十佳武侠游戏”这类模块，顶层 section 没有 name，
    # 真正显示在页面上的标题来自 editorial item 自身。
    if section_kind == "422" and item_display == "Media" and presentation["title"]:
        return presentation["title"]

    if presentation["title"]:
        return presentation["title"]
    if app_title:
        return app_title
    if media_mode == "carousel":
        return "顶部轮播"
    return "未命名区域"


def make_placement(data, options, page_type, page_label, section, item, app_node, event_node, media_mode, placement_type, position, module_position, item_position, group_item_count, content_position, overall_position, path, updated_at, page_date):
    item_attrs = (item or {}).get("attributes") or {}
    presentation = presentation_from_editorial_item(item)
    event = event_details_from_node(event_node, app_node)
    app_title = title_from_node(app_node) or options["gameName"] or ""
    app_subtitle = subtitle_from_node(app_node)
    icon_image = app_icon_from_node(app_node, options.get("appIcon", ""))
    editorial_image = editorial_image_from_node(item) or editorial_image_from_node(section)
    event_image = (event or {}).get("image") or ""
    event_video = (event or {}).get("video") or ""
    is_today_editorial = page_type == "today" and not event
    image = editorial_image or event_image or icon_image if media_mode in ("carousel", "hero", "event") else icon_image
    group_source = presentation_from_editorial_group(section, page_label or "Today") if page_type == "today" else None
    raw_group_title = group_title_from_section(section, page_label if page_type == "today" else "")
    group_title = (group_source["title"] if group_source else "") or raw_group_title or page_label or "Today" if page_type == "today" else raw_group_title or ""
    group_subtitle = ((group_source["subtitle"] if group_source else "") or group_subtitle_from_section(section) or "") if page_type == "today" else group_subtitle_from_section(section) or ""
    today_card_title = presentation["title"] or ((event or {}).get("title")) or app_title or group_title or "展位"
    today_card_subtitle = presentation["subtitle"] or ((event or {}).get("subtitle")) or app_subtitle or ""
    section_title = today_card_title if page_type == "today" else (presentation["title"] or title_from_node(section) or "未命名区域") if is_today_editorial else games_section_title(section, item, app_title, media_mode)
    section_subtitle = today_card_subtitle if page_type == "today" else (presentation["subtitle"] or subtitle_from_node(section)) if is_today_editorial else subtitle_from_node(section)
    default_placement_type = section_title if event or is_today_editorial else "内容列表"
    placement_title = today_card_title if page_type == "today" else (presentation["title"] or ((event or {}).get("title"))) if event else (app_title or presentation["title"] or section_title) if is_today_editorial else (presentation["title"] or app_title or section_title)
    subtitle = today_card_subtitle if page_type == "today" else (presentation["subtitle"] or ((event or {}).get("subtitle"))) if event else (app_subtitle or "") if is_today_editorial else (presentation["subtitle"] or app_subtitle or "")
    description = (event or {}).get("description") or clean_text(node_text_blob(item), 260)
    app_name_text = app_name_text_blob(app_node)
    search_text = normalize("".join(filter(None, [
        app_name_text, group_title, group_subtitle, section_title, section_subtitle,
        placement_title, subtitle, description, (event or {}).get("title"), (event or {}).get("subtitle"),
        (event or {}).get("kind"), (event or {}).get("description"), app_title, app_subtitle, node_text_blob(item)
    ]))) if page_type == "today" else app_name_text
    return {
        "pageType": page_type,
        "pageLabel": page_label,
        "sectionTitle": section_title,
        "sectionSubtitle": section_subtitle,
        "groupTitle": group_title,
        "groupSubtitle": group_subtitle,
        "headerTitle": clean_text(item_attrs.get("headerName") or ""),
        "placementType": placement_type or default_placement_type,
        "placementTitle": placement_title,
        "subtitle": subtitle,
        "description": description,
        "eventStatus": (event or {}).get("status") or "",
        "eventBadgeKind": (event or {}).get("badgeKind") or "",
        "eventKind": (presentation["badge"] or (event or {}).get("kind")) if event else "",
        "eventStartDate": (event or {}).get("startDate") or "",
        "eventEndDate": (event or {}).get("endDate") or "",
        "eventRequirement": (event or {}).get("requirement") or "",
        "callToAction": presentation["callToAction"] or "",
        "appTitle": (event or {}).get("appTitle") or app_title,
        "appSubtitle": (event or {}).get("appSubtitle") or app_subtitle,
        "appIcon": (event or {}).get("appIcon") or icon_image,
        "heroRibbon": (event or {}).get("status") or "",
        "heroEyebrow": presentation["badge"] or ((event or {}).get("kind")) or "",
        "heroTitle": presentation["title"] or ((event or {}).get("title")) or placement_title,
        "heroDescription": presentation["subtitle"] or ((event or {}).get("description")) or subtitle or "",
        "buttonNote": (event or {}).get("requirement") or "",
        "position": position,
        "modulePosition": module_position,
        "itemPosition": item_position,
        "groupItemCount": group_item_count,
        "contentPosition": content_position,
        "overallPosition": overall_position,
        "updatedAt": updated_at or (((section or {}).get("attributes") or {}).get("lastModifiedDate")) or page_date or extract_page_date(data),
        "checkedAt": options["checkedAt"],
        "type": "apps",
        "id": str((app_node or {}).get("id") or (((app_node or {}).get("attributes") or {}).get("adamId")) or ""),
        "appNameText": app_name_text,
        "searchText": search_text,
        "image": image,
        "video": event_video,
        "heroImage": editorial_image or event_image,
        "iconImage": icon_image,
        "mediaMode": media_mode,
        "editorialKind": (((section or {}).get("attributes") or {}).get("editorialElementKind")) or "",
        "path": path,
        "textSnippet": clean_text(node_text_blob(item), 260),
    }


def extract_page_date(data):
    results = (((data or {}).get("results") or {}).get("data")) or []
    if results:
        return results[0].get("date") or ""
    catalog = (data or {}).get("data") or []
    return ((catalog[0].get("attributes") or {}).get("lastModifiedDate")) if catalog else ""


def build_today_placement_index(data, options):
    page = ((((data or {}).get("results") or {}).get("data")) or [{}])[0]
    contents = page.get("contents") or []
    placements = []
    overall_position = 0
    for section_index, section in enumerate(contents):
        recommendations = today_recommendations(section)
        cards = recommendations if recommendations else [{"item": section, "rawIndex": 0}]
        cards = recommendations if recommendations else [section]
        for visual_index, item in enumerate(cards):
            raw_index = visual_index
            parts = direct_placement_parts(item)
            if not parts:
                continue
            overall_position += 1
            for part in parts:
                suffix = f".{part['relationshipKey']}[{part['relatedIndex']}]" if part["relationshipKey"] else ""
                placements.append(make_placement(
                    data=data,
                    options=options,
                    page_type="today",
                    page_label=PAGE_TYPES["today"]["label"],
                    section=section,
                    item=item,
                    app_node=part["appNode"],
                    event_node=part["eventNode"],
                    media_mode="event" if part["eventNode"] else ("hero" if editorial_image_from_node(item) else "icon"),
                    placement_type=title_from_node(section) or PAGE_TYPES["today"]["label"] if part["eventNode"] else source_from_node(section)["title"] or "Today",
                    position=visual_index + 1,
                    module_position=section_index + 1,
                    item_position=visual_index + 1,
                    group_item_count=len(cards),
                    content_position=part["relatedIndex"] + 1 if isinstance(part["relatedIndex"], int) else None,
                    overall_position=overall_position,
                    path=f"today.contents[{section_index}].recommendations[{raw_index}]{suffix}",
                    updated_at=((section.get("attributes") or {}).get("lastModifiedDate")) or page.get("date") or "",
                    page_date=page.get("date") or "",
                ))
    return placements


def games_children(data):
    return (((((((data or {}).get("data") or [{}])[0].get("relationships") or {}).get("tabs") or {}).get("data") or [{}])[0].get("relationships") or {}).get("children") or {}).get("data") or []


def build_games_placement_index(data, options):
    children = games_children(data)
    placements = []
    for section_index, section in enumerate(children):
        kind = str((((section or {}).get("attributes") or {}).get("editorialElementKind")) or "")
        if kind == "415":
            visible_position = 0
            for slide_index, slide in enumerate((((section.get("relationships") or {}).get("children") or {}).get("data")) or []):
                item = (((slide.get("relationships") or {}).get("contents") or {}).get("data") or [None])[0]
                if not item or not editorial_image_from_node(item):
                    continue
                visible_position += 1
                for part in direct_placement_parts(item):
                    slide_section = {"attributes": {**(slide.get("attributes") or {}), "name": "顶部轮播"}}
                    placements.append(make_placement(
                        data=data,
                        options=options,
                        page_type="games",
                        page_label=PAGE_TYPES["games"]["label"],
                        section=slide_section,
                        item=item,
                        app_node=part["appNode"],
                        event_node=part["eventNode"],
                        media_mode="carousel",
                        placement_type="顶部轮播/活动" if part["eventNode"] else "顶部轮播",
                        position=visible_position,
                        module_position=section_index + 1,
                        item_position=visible_position,
                        group_item_count=0,
                        content_position=None,
                        overall_position=None,
                        path=f"games.carousel[{slide_index}]",
                        updated_at=((slide.get("attributes") or {}).get("lastModifiedDate")) or ((section.get("attributes") or {}).get("lastModifiedDate")) or "",
                        page_date="",
                    ))
            continue
        nested_sections = (((section.get("relationships") or {}).get("children") or {}).get("data")) or [section]
        for nested_index, nested in enumerate(nested_sections):
            contents = (((nested.get("relationships") or {}).get("contents") or {}).get("data")) or []
            for item_index, item in enumerate(contents):
                parts = direct_placement_parts(item)
                if not parts:
                    continue
                for part in parts:
                    placements.append(make_placement(
                        data=data,
                        options=options,
                        page_type="games",
                        page_label=PAGE_TYPES["games"]["label"],
                        section=nested,
                        item=item,
                        app_node=part["appNode"],
                        event_node=part["eventNode"],
                        media_mode="event" if part["eventNode"] else "icon",
                        placement_type=title_from_node(nested) or "活动" if part["eventNode"] else "内容列表",
                        position=item_index + 1,
                        module_position=section_index + 1,
                        item_position=item_index + 1,
                        group_item_count=0,
                        content_position=None,
                        overall_position=None,
                        path=f"games.children[{section_index}].{nested_index}.contents[{item_index}]",
                        updated_at=((nested.get("attributes") or {}).get("lastModifiedDate")) or ((section.get("attributes") or {}).get("lastModifiedDate")) or "",
                        page_date="",
                    ))
    return placements


def build_placement_index(data, options):
    return build_today_placement_index(data, options) if options["pageType"] == "today" else build_games_placement_index(data, options)


def analyze_editorial_json(data, options):
    terms = list({normalize(options["gameName"]), *[normalize(alias) for alias in options.get("aliases", []) if alias]})
    terms = [term for term in terms if term]
    app_id = str(options.get("appId") or "").strip()
    placement_index = build_placement_index(data, options)
    exact_id_matches = [
        placement for placement in placement_index
        if app_id and str(placement.get("id") or "").strip() == app_id
    ]
    if app_id:
        indexed_matches = exact_id_matches
    else:
        indexed_matches = [
            placement for placement in placement_index
            if any(term and term in (placement.get("searchText") or placement.get("appNameText") or "") for term in terms)
        ]
    unique = []
    unique_keys = set()
    for match in indexed_matches:
        key = "|".join([
            str(match.get("pageType") or ""),
            str(match.get("mediaMode") or ""),
            str(match.get("id") or ""),
            str(match.get("sectionTitle") or ""),
            str(match.get("modulePosition") or ""),
            str(match.get("itemPosition") or match.get("position") or ""),
            str(match.get("placementTitle") or ""),
        ])
        if key in unique_keys:
            continue
        unique_keys.add(key)
        unique.append(match)
    icon_by_id = {}
    for match in unique:
        if match.get("id") and match.get("appIcon"):
            icon_by_id[match["id"]] = match["appIcon"]
        if match.get("id") and match.get("iconImage"):
            icon_by_id[match["id"]] = match["iconImage"]
    for match in unique:
        fallback_icon = icon_by_id.get(match.get("id") or "", "")
        if not fallback_icon and match.get("id") and not has_placement_image(match):
            fallback_icon = lookup_app_icon(match.get("id"), options.get("countryCode") or "")
        if not match.get("appIcon") and fallback_icon:
            match["appIcon"] = fallback_icon
        if not match.get("iconImage") and fallback_icon:
            match["iconImage"] = fallback_icon
        if match.get("mediaMode") == "icon" and not match.get("image") and fallback_icon:
            match["image"] = fallback_icon
    unique = [match for match in unique if has_placement_image(match)]
    if options["pageType"] == "games":
        rank = {"carousel": 0, "event": 1, "icon": 2}
        unique.sort(key=lambda item: (rank.get(item.get("mediaMode"), 9), item.get("position") or 9999))
    else:
        unique.sort(key=lambda item: (item.get("modulePosition") or 9999, item.get("itemPosition") or item.get("position") or 9999))
    return unique[:80]


def analyze_country(country_code, game_name, app_id, page_types, aliases, checked_at, app_icon):
    country = COUNTRIES[country_code]
    results = []
    for page_type in page_types:
        url = editorial_url(country_code, page_type)
        try:
            data = fetch_json(url, country, retries=2, editorial=True)
            matches = analyze_editorial_json(data, {
                "countryCode": country_code,
                "gameName": game_name,
                "appId": app_id,
                "aliases": aliases,
                "pageType": page_type,
                "checkedAt": checked_at,
                "appIcon": app_icon,
            })
            results.append({
                "country": country_code,
                "countryLabel": country["label"],
                "localName": country["localName"],
                "pageType": page_type,
                "pageLabel": PAGE_TYPES[page_type]["label"],
                "url": url,
                "found": len(matches) > 0,
                "matches": matches,
            })
        except Exception as error:
            results.append({
                "country": country_code,
                "countryLabel": country["label"],
                "localName": country["localName"],
                "pageType": page_type,
                "pageLabel": PAGE_TYPES[page_type]["label"],
                "url": url,
                "found": False,
                "matches": [],
                "error": "Apple API 临时限流，请稍后重新分析" if isinstance(error, AppleCapacityError) or re.search(r"429|capacity exceeded", str(error), re.I) else str(error) or "接口请求失败",
            })
    return results


def analyze_country_page_payload(country_code, page_type, data, game_name, app_id, aliases, checked_at, app_icon):
    country = COUNTRIES[country_code]
    url = editorial_url(country_code, page_type)
    matches = analyze_editorial_json(data, {
        "countryCode": country_code,
        "gameName": game_name,
        "appId": app_id,
        "aliases": aliases,
        "pageType": page_type,
        "checkedAt": checked_at,
        "appIcon": app_icon,
    })
    return {
        "country": country_code,
        "countryLabel": country["label"],
        "localName": country["localName"],
        "pageType": page_type,
        "pageLabel": PAGE_TYPES[page_type]["label"],
        "url": url,
        "found": len(matches) > 0,
        "matches": matches,
    }


class AppHandler(BaseHTTPRequestHandler):
    server_version = "AppExpoPython/1.0"

    def do_GET(self):
        try:
            parsed = urllib.parse.urlparse(self.path)
            if parsed.path == "/api/config":
                self.json_response(200, {
                    "countries": [{"code": code, **country} for code, country in COUNTRIES.items()],
                    "pageTypes": [{"value": value, "label": page["label"]} for value, page in PAGE_TYPES.items()],
                })
                return
            if parsed.path == "/api/games":
                country_code = (urllib.parse.parse_qs(parsed.query).get("country") or ["CN"])[0].upper()
                if country_code not in COUNTRIES:
                    self.json_response(400, {"error": "不支持的国家"})
                    return
                self.json_response(200, static_developer_games(country_code))
                return
            if parsed.path == "/api/game-catalog":
                self.json_response(200, global_game_catalog())
                return
            if parsed.path == "/api/sync-config":
                self.json_response(200, sync_target_config())
                return
            if parsed.path == "/api/scheduler":
                self.json_response(200, {
                    "scheduler": scheduler_snapshot(),
                    "syncTarget": sync_target_config(),
                })
                return
            if parsed.path == "/api/weekly-games":
                query = urllib.parse.parse_qs(parsed.query)
                country_code = clean_text((query.get("country") or ["CN"])[0], 20).upper()
                week_start = clean_text((query.get("weekStart") or [""])[0], 20)
                self.json_response(200, load_weekly_games(country_code=country_code, week_start=week_start))
                return
            if parsed.path == "/api/weekly-games/weeks":
                query = urllib.parse.parse_qs(parsed.query)
                country_code = clean_text((query.get("country") or ["CN"])[0], 20).upper()
                self.json_response(200, load_weekly_weeks(country_code=country_code))
                return
            if parsed.path == "/api/weekly-games/status":
                self.json_response(200, {
                    "scheduler": weekly_snapshot(),
                    "todayCapturedCountries": weekly_capture_count(local_today()),
                    "countryCount": len(COUNTRIES),
                })
                return
            if parsed.path == "/api/history/runs":
                query = urllib.parse.parse_qs(parsed.query)
                page_num = int((query.get("page") or ["1"])[0])
                page_size = int((query.get("size") or ["20"])[0])
                keyword = clean_text((query.get("keyword") or [""])[0], 120)
                game_key = clean_text((query.get("gameKey") or [""])[0], 160)
                sync_date = clean_text((query.get("date") or [""])[0], 20)
                self.json_response(200, load_history_runs(page_num=page_num, page_size=page_size, keyword=keyword, sync_date=sync_date, game_key=game_key))
                return
            if parsed.path == "/api/history/games":
                self.json_response(200, load_history_games())
                return
            if parsed.path == "/api/history/run-games":
                query = urllib.parse.parse_qs(parsed.query)
                run_id = clean_text((query.get("runId") or [""])[0], 160)
                self.json_response(200, load_history_run_games(run_id))
                return
            if parsed.path == "/api/history/countries":
                query = urllib.parse.parse_qs(parsed.query)
                game_key = clean_text((query.get("gameKey") or [""])[0], 160)
                app_id = clean_text((query.get("appId") or [""])[0], 80)
                page_type = clean_text((query.get("pageType") or [""])[0], 20)
                date_from = clean_text((query.get("dateFrom") or [""])[0], 20)
                date_to = clean_text((query.get("dateTo") or [""])[0], 20)
                self.json_response(200, load_history_countries(game_key=game_key, app_id=app_id, date_from=date_from, date_to=date_to, page_type=page_type))
                return
            if parsed.path == "/api/history/analytics":
                query = urllib.parse.parse_qs(parsed.query)
                game_key = clean_text((query.get("gameKey") or [""])[0], 160)
                app_id = clean_text((query.get("appId") or [""])[0], 80)
                country = clean_text((query.get("country") or [""])[0], 20)
                page_type = clean_text((query.get("pageType") or [""])[0], 20)
                date_from = clean_text((query.get("dateFrom") or [""])[0], 20)
                date_to = clean_text((query.get("dateTo") or [""])[0], 20)
                self.json_response(200, load_history_analytics(
                    game_key=game_key,
                    app_id=app_id,
                    country=country,
                    page_type=page_type,
                    date_from=date_from,
                    date_to=date_to,
                ))
                return
            if parsed.path == "/api/history/matches":
                query = urllib.parse.parse_qs(parsed.query)
                page_num = int((query.get("page") or ["1"])[0])
                page_size = int((query.get("size") or ["100"])[0])
                run_id = clean_text((query.get("runId") or [""])[0], 160)
                game_key = clean_text((query.get("gameKey") or [""])[0], 160)
                country = clean_text((query.get("country") or [""])[0], 20)
                page_type = clean_text((query.get("pageType") or [""])[0], 20)
                sync_date = clean_text((query.get("date") or [""])[0], 20)
                self.json_response(200, load_history_matches(
                    run_id=run_id,
                    game_key=game_key,
                    page_num=page_num,
                    page_size=page_size,
                    country=country,
                    page_type=page_type,
                    sync_date=sync_date,
                ))
                return
            if parsed.path == "/api/image-data":
                image_url = (urllib.parse.parse_qs(parsed.query).get("url") or [""])[0]
                try:
                    self.json_response(200, {"dataUrl": fetch_image_data_url(image_url)})
                except Exception as error:
                    self.json_response(400, {"error": str(error) or "图片转换失败"})
                return
            self.serve_static(parsed.path)
        except Exception as error:
            self.json_response(500, {"error": str(error) or "服务器错误"})

    def do_POST(self):
        try:
            parsed = urllib.parse.urlparse(self.path)
            body = self.parse_body()
            if parsed.path == "/api/export-result":
                result = body.get("result") or {}
                filename = safe_ascii_download_name(
                    f"{result.get('country') or 'AppExpo'}-{result.get('pageLabel') or result.get('pageType') or 'placement'}-{result.get('localName') or ''}"
                )
                buffer = run_python_export({
                    "result": result,
                    "checkedAt": body.get("checkedAt") or "",
                    "theme": "dark" if body.get("theme") == "dark" else "light",
                })
                self.send_response(200)
                self.send_header("content-type", "image/png")
                self.send_header("content-disposition", content_disposition(f"{filename}.png"))
                self.send_header("x-export-filename", f"{safe_ascii_download_name(filename)}.png")
                self.send_header("cache-control", "no-store")
                self.end_headers()
                self.wfile.write(buffer)
                return
            if parsed.path == "/api/analyze":
                countries = [str(item).upper() for item in body.get("countries", [body.get("country", "CN")])]
                page_types = [str(item).lower() for item in body.get("pageTypes", ["today", "games"]) if str(item).lower() in PAGE_TYPES]
                game_name = clean_text(body.get("gameName"), 120)
                app_id = str(body.get("appId") or "")
                app_icon = str(body.get("appIcon") or "")
                aliases = [clean_text(item, 120) for item in body.get("aliases", []) if clean_text(item, 120)]
                country_entries = body.get("countryEntries") or {}
                if not game_name:
                    self.json_response(400, {"error": "请先选择游戏"})
                    return
                if not countries or any(country not in COUNTRIES for country in countries):
                    self.json_response(400, {"error": "不支持的国家"})
                    return
                if not page_types:
                    self.json_response(400, {"error": "请至少选择 Today 或 Games"})
                    return
                checked_at = datetime.now(timezone.utc).isoformat().replace("+00:00", "Z")
                results = []
                for country_code in countries:
                    country_entry = country_entries.get(country_code) if isinstance(country_entries, dict) else None
                    results.extend(analyze_country(
                        country_code,
                        clean_text((country_entry or {}).get("name"), 120) or game_name,
                        str((country_entry or {}).get("appId") or app_id),
                        page_types,
                        [clean_text(item, 120) for item in ((country_entry or {}).get("aliases") or aliases) if clean_text(item, 120)],
                        checked_at,
                        str((country_entry or {}).get("icon") or app_icon),
                    ))
                self.json_response(200, {
                    "gameName": game_name,
                    "appId": app_id,
                    "countries": countries,
                    "pageTypes": page_types,
                    "checkedAt": checked_at,
                    "crossCountry": bool(body.get("crossCountry")),
                    "results": results,
                })
                return
            if parsed.path == "/api/history/analytics-batch":
                date_from = clean_text(body.get("dateFrom") or "", 20)
                date_to = clean_text(body.get("dateTo") or "", 20)
                page_type = clean_text(body.get("pageType") or "", 20)
                selections = body.get("selections") or []
                if date_from and date_to and date_from > date_to:
                    self.json_response(400, {"error": "开始日期不能晚于结束日期"})
                    return
                if page_type and page_type not in PAGE_TYPES:
                    self.json_response(400, {"error": "不支持的页面类型"})
                    return
                self.json_response(200, load_history_analytics_batch(
                    selections=selections,
                    date_from=date_from,
                    date_to=date_to,
                    page_type=page_type,
                ))
                return
            if parsed.path == "/api/sync-preview":
                self.json_response(200, build_sync_payload(body))
                return
            if parsed.path == "/api/sync-deliver":
                payload = body.get("payload") or build_sync_payload(body)
                delivery = post_sync_payload(payload)
                self.json_response(200, {
                    "ok": True,
                    "target": sync_target_config(),
                    "delivery": delivery,
                })
                return
            if parsed.path == "/api/scheduler/start":
                interval_minutes = int(body.get("intervalMinutes") or 120)
                self.json_response(200, {
                    "ok": True,
                    "scheduler": start_scheduler(interval_minutes),
                    "syncTarget": sync_target_config(),
                })
                return
            if parsed.path == "/api/scheduler/stop":
                self.json_response(200, {
                    "ok": True,
                    "scheduler": stop_scheduler(),
                    "syncTarget": sync_target_config(),
                })
                return
            if parsed.path == "/api/scheduler/run-once":
                snapshot = scheduler_snapshot()
                if snapshot["running"]:
                    self.json_response(400, {"error": "当前已有定时任务正在执行"})
                    return
                ensure_scheduler_thread()
                append_scheduler_trace("scheduler_run_once_requested")
                update_scheduler_state(
                    running=True,
                    lastRunAt=utc_now_iso(),
                    lastFinishedAt="",
                    lastStatus="running",
                    lastError="",
                    currentGame="准备中",
                    currentIndex=0,
                    totalGames=0,
                    currentTriggerSource="manual-run-once",
                    lastTriggerSource="manual-run-once",
                )
                threading.Thread(
                    target=run_scheduled_full_sync,
                    args=("manual-run-once",),
                    daemon=True,
                    name="appexpo-scheduler-once",
                ).start()
                self.json_response(200, {
                    "ok": True,
                    "scheduler": scheduler_snapshot(),
                    "syncTarget": sync_target_config(),
                })
                return
            if parsed.path == "/api/weekly-games/capture":
                snapshot = weekly_snapshot()
                if snapshot["running"]:
                    self.json_response(400, {"error": "每周一更抓取正在执行"})
                    return
                countries = body.get("countries") or []
                if not countries and body.get("country"):
                    countries = [body.get("country")]
                country_codes = [str(code).upper() for code in countries if str(code).upper() in COUNTRIES] or sorted(COUNTRIES.keys())
                force = bool(body.get("force"))
                threading.Thread(
                    target=capture_weekly_games_for_countries,
                    args=(country_codes, force, "manual"),
                    daemon=True,
                    name="appexpo-weekly-games-manual",
                ).start()
                self.json_response(200, {
                    "ok": True,
                    "scheduler": weekly_snapshot(),
                    "countryCount": len(country_codes),
                })
                return
            self.json_response(405, {"error": "Method not allowed"})
        except SyncUploadError as error:
            self.json_response(400, {"error": str(error) or "上传失败"})
        except Exception as error:
            self.json_response(500, {"error": str(error) or "服务器错误"})

    def parse_body(self):
        length = int(self.headers.get("content-length") or "0")
        if length > 1_000_000:
            raise RuntimeError("请求体过大")
        raw = self.rfile.read(length) if length else b"{}"
        try:
            return json.loads(raw.decode("utf-8"))
        except Exception:
            raise RuntimeError("请求 JSON 格式不正确")

    def serve_static(self, request_path):
        relative = "/index.html" if request_path == "/" else urllib.parse.unquote(request_path)
        safe = os.path.normpath(relative).lstrip(os.sep)
        file_path = (PUBLIC_DIR / safe).resolve()
        if not str(file_path).startswith(str(PUBLIC_DIR.resolve())):
            self.send_response(403)
            self.end_headers()
            self.wfile.write(b"Forbidden")
            return
        if not file_path.exists() or not file_path.is_file():
            self.send_response(404)
            self.send_header("content-type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write(b"Not found")
            return
        content = file_path.read_bytes()
        mime = MIME_TYPES.get(file_path.suffix.lower()) or mimetypes.guess_type(str(file_path))[0] or "application/octet-stream"
        self.send_response(200)
        self.send_header("content-type", mime)
        self.send_header("cache-control", "no-store")
        self.end_headers()
        self.wfile.write(content)

    def json_response(self, status, payload):
        raw = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("content-type", "application/json; charset=utf-8")
        self.send_header("cache-control", "no-store")
        self.end_headers()
        self.wfile.write(raw)

    def log_message(self, format, *args):
        return


def main():
    init_local_db()
    ensure_weekly_scheduler_thread()
    server = ThreadingHTTPServer(("0.0.0.0", PORT), AppHandler)
    print(f"Apple Store placement analyzer running at http://localhost:{PORT}")
    server.serve_forever()


if __name__ == "__main__":
    main()
